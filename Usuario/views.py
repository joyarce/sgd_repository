#Usuario\views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db import connection, transaction

#from plantillas_documentos_tecnicos.utils_documentos import extract_blob_name_from_signed_url
from plantillas_documentos_tecnicos.utils_documentos import (
    extract_blob_name_from_gcs_path,
    extract_blob_name_from_signed_url,
    inicializar_version_inicial,
    insertar_documento_generado
)


# === Utilidades y librerías externas ===
from google.cloud import storage
from django.conf import settings
from collections import Counter, defaultdict
from datetime import datetime, timedelta
import json, re, unidecode, csv, openpyxl, traceback

# === Modelos locales ===
from .models import FilePreview

# === Librerías especializadas ===
from openpyxl import load_workbook


import re



def to_int_or_none(value):
    try:
        return int(value)
    except:
        return None



def clean(x):
    """Limpia nombres para rutas GCS usando un solo guion bajo por cada token no válido."""
    x = re.sub(r"[\/\\]+", " ", x)
    x = re.sub(r"\s+", "_", x)
    x = re.sub(r"[:*?\"<>|]+", "_", x)
    x = re.sub(r"_+", "_", x)
    return x.strip("_")


# Duración de previsualización en minutos
PREVIEW_EXPIRATION_MINUTES = 60  

# Tipos MIME permitidos (puedes ampliar)
ALLOWED_MIME_TYPES = [
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/pdf",
    "image/jpeg",
    "image/png",
]

# Cliente de Google Cloud Storage
storage_client = storage.Client.from_service_account_json(settings.GCP_SERVICE_ACCOUNT_JSON)
bucket = storage_client.bucket(settings.GCP_BUCKET_NAME)


def get_or_create_preview_url(blob):
    """Genera o reutiliza enlace de previsualización temporal"""
    try:
        preview = FilePreview.objects.get(blob_name=blob.name)
        if preview.is_expired():
            raise FilePreview.DoesNotExist
        remaining_minutes = int((preview.expires_at - timezone.now()).total_seconds() / 60)
        return preview.signed_url, remaining_minutes
    except FilePreview.DoesNotExist:
        try:
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=PREVIEW_EXPIRATION_MINUTES),
                method="GET"
            )
        except Exception:
            url = ""
        expires_at = timezone.now() + timedelta(minutes=PREVIEW_EXPIRATION_MINUTES)
        FilePreview.objects.update_or_create(
            blob_name=blob.name,
            defaults={"signed_url": url, "expires_at": expires_at}
        )
        return url, PREVIEW_EXPIRATION_MINUTES

@login_required
def inicio(request):
    return render(request, "usuario_inicio.html")

@login_required
def lista_proyectos(request):
    proyectos = []

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, nombre, descripcion, 
                   fecha_recepcion_evaluacion, 
                   fecha_inicio_planificacion,
                   fecha_inicio_ejecucion,
                   fecha_cierre_proyecto
            FROM proyectos
            ORDER BY nombre
        """)
        rows = cursor.fetchall()
        for row in rows:
            proyectos.append({
                "id": row[0],
                "nombre": row[1],
                "descripcion": row[2],
                "fecha_recepcion_evaluacion": row[3],
                "fecha_inicio_planificacion": row[4],
                "fecha_inicio_ejecucion": row[5],
                "fecha_cierre_proyecto": row[6],
            })

    return render(request, "usuario_proyectos.html", {"proyectos": proyectos})



@require_POST
def validar_orden_ajax(request):
    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES.get("archivo")

        if not archivo.name.lower().endswith(".xlsx"):
            return JsonResponse({"error": "Formato de archivo no soportado. Se espera un .xlsx"}, status=400)

        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            numero_servicio = None

            for defined_name in wb.defined_names.values():
                if defined_name.name.lower() == "numordenservicio":
                    dest = list(defined_name.destinations)[0]
                    sheet_name, cell_coord = dest
                    sheet = wb[sheet_name]
                    valor_celda = sheet[cell_coord].value

                    if valor_celda is not None:
                        numero_servicio = str(valor_celda).strip()
                        break

            if numero_servicio:
                return JsonResponse({"numero_servicio": numero_servicio})
            else:
                return JsonResponse({"error": "No se encontró el nombre definido 'NumOrdenServicio' o la celda está vacía."}, status=400)

        except Exception as e:
            return JsonResponse({"error": f"Error al procesar el archivo: {str(e)}"}, status=500)

    return JsonResponse({"error": "Petición inválida o falta el archivo."}, status=400)

@login_required
def detalle_proyecto(request, proyecto_id):
    """
    Muestra los detalles completos de un proyecto incluyendo:
    - Proyecto
    - Contrato
    - Cliente (con abreviatura)
    - Máquinas (usa abreviatura)
    - Requerimientos (incluye abreviatura del documento técnico y CONFIDENCIALIDAD)
    """

    sql = """
    WITH UltimoEstado AS (
        SELECT
            requerimiento_id,
            estado_destino_id,
            ROW_NUMBER() OVER (
                PARTITION BY requerimiento_id 
                ORDER BY fecha_cambio DESC, id DESC
            ) AS rn
        FROM public.log_estado_requerimiento_documento
    )
    SELECT
        -- Proyecto
        P.id AS proyecto_id,
        P.nombre AS nombre_proyecto,
        P.descripcion AS proyecto_descripcion,
        P.numero_servicio,
        P.fecha_recepcion_evaluacion,
        P.fecha_inicio_planificacion,
        P.fecha_inicio_ejecucion,
        P.fecha_cierre_proyecto,
        P.abreviatura AS proyecto_abreviatura,
        P.path_gcs,

        -- Administrador
        U.nombre AS administrador_nombre_completo,
        U.email AS administrador_email,

        -- Faena
        F.nombre AS nombre_faena,

        -- Contrato
        C.id AS contrato_id,
        C.numero_contrato,
        C.monto_total,
        C.fecha_firma AS contrato_fecha_firma,
        C.representante_cliente_nombre,
        C.representante_cliente_correo,
        C.representante_cliente_telefono,

        -- Cliente
        CL.id AS cliente_id,
        CL.nombre AS cliente_nombre,
        CL.abreviatura AS cliente_abreviatura,
        CL.rut AS cliente_rut,
        CL.direccion AS cliente_direccion,
        CL.correo_contacto AS cliente_correo,
        CL.telefono_contacto AS cliente_telefono,

        -- Requerimientos
        RDT.id AS requerimiento_id,
        TDT.nombre AS nombre_documento_tecnico,
        TDT.abreviatura AS abreviatura_documento_tecnico,
        E.nombre AS estado_actual_documento,
        RDT.fecha_registro AS requerimiento_fecha,
        RDT.confidencialidad AS nivel_confidencialidad,

        -- Máquinas
        M.id AS maquina_id,
        M.nombre AS maquina_nombre,
        M.abreviatura,
        M.marca,
        M.modelo,
        M.anio_fabricacion,
        M.tipo AS tipo_maquina,
        M.descripcion AS maquina_descripcion

    FROM public.proyectos P
    INNER JOIN public.usuarios U ON P.administrador_id = U.id
    LEFT JOIN public.faenas F ON P.faena_id = F.id
    LEFT JOIN public.contratos C ON P.contrato_id = C.id
    LEFT JOIN public.clientes CL ON C.cliente_id = CL.id
    LEFT JOIN public.requerimiento_documento_tecnico RDT ON P.id = RDT.proyecto_id
    LEFT JOIN public.tipo_documentos_tecnicos TDT ON RDT.tipo_documento_id = TDT.id
    LEFT JOIN UltimoEstado UE ON RDT.id = UE.requerimiento_id AND UE.rn = 1
    LEFT JOIN public.estado_documento E ON UE.estado_destino_id = E.id
    LEFT JOIN public.proyecto_maquina PM ON P.id = PM.proyecto_id
    LEFT JOIN public.maquinas M ON PM.maquina_id = M.id
    WHERE P.id = %s
    ORDER BY RDT.fecha_registro DESC, M.id;
    """

    with connection.cursor() as cursor:
        cursor.execute(sql, [proyecto_id])
        columns = [col[0] for col in cursor.description]
        resultados = [dict(zip(columns, row)) for row in cursor.fetchall()]

    proyecto_info = resultados[0] if resultados else {}
    requerimientos, maquinas = [], []
    req_ids, maquina_ids = set(), set()

    for row in resultados:

        # Agregar requerimientos únicos
        if row['requerimiento_id'] and row['requerimiento_id'] not in req_ids:
            requerimientos.append({
                'id': row['requerimiento_id'],
                'nombre_documento_tecnico': row['nombre_documento_tecnico'],
                'abreviatura': row['abreviatura_documento_tecnico'],
                'estado_actual': row['estado_actual_documento'],
                'fecha_registro': row['requerimiento_fecha'],
                'confidencialidad': row['nivel_confidencialidad'],
            })
            req_ids.add(row['requerimiento_id'])

        # Agregar máquinas únicas
        if row['maquina_id'] and row['maquina_id'] not in maquina_ids:
            maquinas.append({
                'id': row['maquina_id'],
                'nombre': row['maquina_nombre'],
                'abreviatura': row['abreviatura'],
                'marca': row['marca'],
                'modelo': row['modelo'],
                'anio_fabricacion': row['anio_fabricacion'],
                'tipo': row['tipo_maquina'],
                'descripcion': row['maquina_descripcion'],
            })
            maquina_ids.add(row['maquina_id'])

    return render(request, "usuario_proyecto_detalle.html", {
        'proyecto': proyecto_info,
        'requerimientos': requerimientos,
        'maquinas': maquinas,
    })


@login_required
def eliminar_proyecto(request, proyecto_id):
    if request.method != "POST":
        return redirect("usuario:lista_proyectos")

    try:
        # ============================================================
        # 1️⃣ Obtener datos del proyecto con abreviaturas correctas
        # ============================================================
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    P.abreviatura AS proyecto_abrev,
                    CL.abreviatura AS cliente_abrev
                FROM proyectos P
                JOIN contratos C ON P.contrato_id = C.id
                JOIN clientes CL ON C.cliente_id = CL.id
                WHERE P.id = %s
            """, [proyecto_id])

            row = cursor.fetchone()

        if not row:
            messages.error(request, "❌ Proyecto no encontrado.")
            return redirect("usuario:lista_proyectos")

        proyecto_abrev, cliente_abrev = row

        base_folder = f"DocumentosProyectos/{cliente_abrev}/{proyecto_abrev}/"

        # ============================================================
        # 2️⃣ Obtener todos los signed_url del proyecto
        # ============================================================
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT VDT.signed_url
                FROM version_documento_tecnico VDT
                JOIN requerimiento_documento_tecnico R
                      ON VDT.requerimiento_documento_id = R.id
                WHERE R.proyecto_id = %s
            """, [proyecto_id])

            signed_urls = [r[0] for r in cursor.fetchall() if r[0]]

        storage_client = storage.Client()
        bucket = storage_client.bucket(settings.GCP_BUCKET_NAME)

        # ============================================================
        # 3️⃣ Eliminar archivos individuales por signed_url
        # ============================================================
        for url in signed_urls:
            blob_name = extract_blob_name_from_signed_url(url)
            if blob_name:
                blob = bucket.blob(blob_name)
                if blob.exists():
                    blob.delete()

        # ============================================================
        # 4️⃣ Eliminar carpeta del proyecto completa
        # ============================================================
        for blob in bucket.list_blobs(prefix=base_folder):
            blob.delete()

        # ============================================================
        # 5️⃣ Eliminar datos en BD (en orden correcto)
        # ============================================================
        with transaction.atomic():
            with connection.cursor() as cursor:

                # IDs de requerimientos
                cursor.execute("""
                    SELECT id
                    FROM requerimiento_documento_tecnico
                    WHERE proyecto_id = %s
                """, [proyecto_id])

                reqs = [r[0] for r in cursor.fetchall()]

                if reqs:
                    cursor.execute("""
                        DELETE FROM version_documento_tecnico
                        WHERE requerimiento_documento_id = ANY(%s)
                    """, [reqs])

                    cursor.execute("""
                        DELETE FROM hitos_requerimiento_documento
                        WHERE requerimiento_id = ANY(%s)
                    """, [reqs])

                    cursor.execute("""
                        DELETE FROM requerimiento_equipo_rol
                        WHERE requerimiento_id = ANY(%s)
                    """, [reqs])

                    cursor.execute("""
                        DELETE FROM requerimiento_documento_tecnico
                        WHERE id = ANY(%s)
                    """, [reqs])

                # borrar máquinas del proyecto
                cursor.execute("""
                    DELETE FROM proyecto_maquina
                    WHERE proyecto_id = %s
                """, [proyecto_id])

                # borrar documentos generados
                cursor.execute("""
                    DELETE FROM documentos_generados
                    WHERE proyecto_id = %s
                """, [proyecto_id])

                # borrar proyecto
                cursor.execute("""
                    DELETE FROM proyectos
                    WHERE id = %s
                """, [proyecto_id])

        messages.success(request, "🗑️ Proyecto eliminado correctamente (BD + GCS).")
        return redirect("usuario:lista_proyectos")

    except Exception as e:
        print("⚠️ ERROR eliminar_proyecto:", e)
        messages.error(request, f"Error al eliminar proyecto: {e}")
        return redirect("usuario:lista_proyectos")








@login_required
def detalle_documento(request, documento_id):
    from collections import Counter, defaultdict
    from django.utils import timezone

    logs = []
    equipo_redactores = []
    equipo_revisores = []
    equipo_aprobadores = []
    documento_info = None

    with connection.cursor() as cursor:

        # ============================================================
        # 📌 1. Información principal del documento (tipo + categoría)
        # ============================================================
        cursor.execute("""
            SELECT RDT.id, TDT.nombre, CDT.nombre, RDT.fecha_registro
            FROM requerimiento_documento_tecnico RDT
            LEFT JOIN tipo_documentos_tecnicos TDT ON RDT.tipo_documento_id = TDT.id
            LEFT JOIN categoria_documentos_tecnicos CDT ON TDT.categoria_id = CDT.id
            WHERE RDT.id = %s
        """, [documento_id])

        row = cursor.fetchone()
        if row:
            documento_info = {
                'id': row[0],
                'tipo_documento': row[1],
                'categoria': row[2],
                'created_at': row[3] or timezone.now()
            }

        # ============================================================
        # 📌 2. Logs del documento
        # ============================================================
        cursor.execute("""
            SELECT 
                rdt.id,
                u.nombre AS usuario_nombre,
                rol.nombre AS rol_usuario,
                eo.nombre AS estado_origen,
                ed.nombre AS estado_destino,
                led.created_at AS fecha_accion,
                led.observaciones
            FROM log_estado_requerimiento_documento led
            LEFT JOIN requerimiento_documento_tecnico rdt 
                ON led.requerimiento_id = rdt.id
            LEFT JOIN usuarios u 
                ON led.usuario_id = u.id
            LEFT JOIN requerimiento_equipo_rol rer 
                ON rer.requerimiento_id = rdt.id 
               AND rer.usuario_id = u.id
            LEFT JOIN roles_ciclodocumento rol 
                ON rer.rol_id = rol.id
            LEFT JOIN estado_documento eo 
                ON led.estado_origen_id = eo.id
            LEFT JOIN estado_documento ed 
                ON led.estado_destino_id = ed.id
            WHERE rdt.id = %s
            ORDER BY led.created_at ASC
        """, [documento_id])

        columns = [c[0] for c in cursor.description]
        logs = [dict(zip(columns, r)) for r in cursor.fetchall()]

        # ============================================================
        # 📌 3. Si no tiene log inicial → agregamos uno
        # ============================================================
        if not logs or logs[0]['estado_origen'] is None:
            logs.insert(0, {
                'fecha_accion': documento_info.get("created_at"),
                'estado_origen': 'Sistema',
                'estado_destino': 'Pendiente de Inicio',
                'usuario_nombre': 'Sistema',
                'rol_usuario': '',
                'observaciones': 'El sistema ha habilitado la plantilla en el repositorio.'
            })

        # ============================================================
        # 📌 4. Reconstruir estado_origen cuando es NULL
        # ============================================================
        estado_anterior = "Pendiente de Inicio"
        for log in logs:
            if not log['estado_origen']:
                log['estado_origen'] = estado_anterior
            if not log['usuario_nombre']:
                log['usuario_nombre'] = "Sistema"
            if not log['rol_usuario']:
                log['rol_usuario'] = ""
            estado_anterior = log['estado_destino']

        # ============================================================
        # 📌 5. Equipos (Redactor, Revisor, Aprobador)
        # ============================================================
        cursor.execute("""
            SELECT u.nombre, rol.nombre
            FROM requerimiento_equipo_rol rer
            INNER JOIN usuarios u ON rer.usuario_id = u.id
            INNER JOIN roles_ciclodocumento rol ON rer.rol_id = rol.id
            WHERE rer.requerimiento_id = %s AND rer.activo = true
        """, [documento_id])

        for usuario_nombre, rol_nombre in cursor.fetchall():
            rol = rol_nombre.lower().strip()
            if rol == "redactor":
                equipo_redactores.append(usuario_nombre)
            elif rol == "revisor":
                equipo_revisores.append(usuario_nombre)
            elif rol == "aprobador":
                equipo_aprobadores.append(usuario_nombre)

    # ============================================================
    # 📌 6. Métricas adicionales (acciones, estados, tiempos)
    # ============================================================
    acciones_por_usuario = Counter(log['usuario_nombre'] for log in logs)

    conteo_estados = Counter(log['estado_destino'] for log in logs)

    # Tiempos en cada estado
    tiempos_estado = defaultdict(list)
    for i in range(1, len(logs)):
        estado_anterior = logs[i - 1]['estado_destino']
        delta = logs[i]['fecha_accion'] - logs[i - 1]['fecha_accion']
        tiempos_estado[estado_anterior].append(delta.total_seconds())

    tiempo_promedio_estado = {
        estado: (sum(tiempos) / len(tiempos))
        for estado, tiempos in tiempos_estado.items()
        if tiempos
    }

    # ============================================================
    # 📌 7. Contexto final
    # ============================================================
    context = {
        "documento": {
            "titulo": f"{documento_info['tipo_documento']} - {documento_info['categoria']}",
            "categoria": documento_info['categoria'],
            "tipo_documento": documento_info['tipo_documento'],
        },
        "equipo_redactores": sorted(equipo_redactores),
        "equipo_revisores": sorted(equipo_revisores),
        "equipo_aprobadores": sorted(equipo_aprobadores),
        "logs": logs,
        "acciones_por_usuario": list(acciones_por_usuario.items()),
        "conteo_estados": dict(conteo_estados),
        "tiempo_promedio_estado": tiempo_promedio_estado,
    }

    return render(request, "usuario_proyecto_detalle_documento.html", context)




@login_required
def lista_usuarios(request):
    usuarios = []

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                u.id,
                u.nombre,
                u.email AS email_corporativo,
                COALESCE(u.email_secundario, '') AS email_secundario,
                COALESCE(u.telefono_corporativo, '') AS telefono_corporativo,
                COALESCE(u.telefono_secundario, '') AS telefono_secundario,
                COALESCE(a.codigo, '') AS area_trabajo,
                COALESCE(c.nombre, '') AS cargo,
                u.fecha_registro
            FROM public.usuarios u
            LEFT JOIN public.cargos_empresa c ON u.cargo_id = c.id
            LEFT JOIN public.area_cargo_empresa a ON c.area_id = a.id
            ORDER BY u.id;
        """)
        rows = cursor.fetchall()
        for row in rows:
            usuarios.append({
                "id": row[0],
                "nombre": row[1],
                "email_corporativo": row[2],
                "email_secundario": row[3],
                "telefono_corporativo": row[4],
                "telefono_secundario": row[5],
                "area_trabajo": row[6],
                "cargo": row[7],
                "fecha_registro": row[8],  # datetime o None
            })

    return render(request, "usuario_usuarios.html", {"usuarios": usuarios})



def pretty_name(name: str) -> str:
    """
    Convierte nombres de carpetas del bucket en algo legible.
    """
    if not name:
        return ""

    if name.startswith("RQ-"):
        return name

    # Underscore → espacio
    name = name.replace("_", " ")

    # Capitalizar cada palabra sin alterar números
    return " ".join(w.capitalize() for w in name.split())



@login_required
def list_files(request):

    ROOTS = ["DocumentosProyectos", "Plantillas"]

    # Carpeta actual solicitada
    raw_folder = request.GET.get("folder", "").strip("/")

    # ============================================================
    # 1) Si NO se especifica carpeta → mostrar las carpetas raíz
    # ============================================================
    if raw_folder == "":
        folders = [{"id": r, "name": r} for r in ROOTS]

        return render(request, "usuario_repositorio.html", {
            "folders": folders,
            "files": [],
            "current_folder": "",
            "parent_folder": "",
            "breadcrumb": [("Repositorio", "")]
        })

    # ============================================================
    # 2) Validación: Debe empezar por DocumentosProyectos o Plantillas
    # ============================================================
    if not any(raw_folder.startswith(r) for r in ROOTS):
        return redirect("usuario:list_files")

    full_path = raw_folder
    prefix = f"{full_path.rstrip('/')}/"

    folders, files = [], []

    try:
        iterator = bucket.list_blobs(prefix=prefix, delimiter="/")
        page = next(iterator.pages)

        # --------------------------------------------------------
        # SUBCARPETAS
        # --------------------------------------------------------
        for p in page.prefixes:

            clean_id = p.strip("/")
            last_part = clean_id.split("/")[-1]

            folders.append({
                "id": clean_id,
                "name": pretty_name(last_part),
                "raw": p
            })

        # --------------------------------------------------------
        # ARCHIVOS
        # --------------------------------------------------------
        for blob in page:
            if blob.name.endswith("/"):
                continue

            preview_url, remaining = get_or_create_preview_url(blob)

            files.append({
                "id": blob.name,
                "name": pretty_name(blob.name.split("/")[-1]),
                "raw_name": blob.name.split("/")[-1],
                "preview_url": preview_url,
                "size": blob.size,
                "created_at": blob.time_created,
            })

    except Exception as e:
        print("ERROR list_files:", e)

    # ============================================================
    # 3) Breadcrumb
    # ============================================================
    breadcrumb = [("Repositorio", "")]
    partes = full_path.split("/")

    accum = ""
    for p in partes:
        accum = p if accum == "" else f"{accum}/{p}"
        breadcrumb.append((pretty_name(p), accum))

    # ============================================================
    # 4) Carpeta padre
    # ============================================================
    parent_parts = partes[:-1]
    parent_folder = "/".join(parent_parts) if parent_parts else ""

    return render(request, "usuario_repositorio.html", {
        "folders": folders,
        "files": files,
        "current_folder": full_path,
        "parent_folder": parent_folder,
        "breadcrumb": breadcrumb,
    })






@login_required
def upload_file(request):
    if request.method == "POST" and request.FILES.get("file"):
        uploaded_file = request.FILES["file"]
        folder = request.POST.get("current_folder", "")

        if uploaded_file.content_type not in ALLOWED_MIME_TYPES:
            return HttpResponse(
                "<script>alert('Tipo de archivo no permitido.'); window.history.back();</script>"
            )

        blob_name = f"{folder}{uploaded_file.name}" if folder else uploaded_file.name
        try:
            blob = bucket.blob(blob_name)
            blob.upload_from_file(uploaded_file, content_type=uploaded_file.content_type)
        except Exception as e:
            return HttpResponse(f"Error al subir archivo: {e}")

        return redirect(f"{request.path}?folder={folder}")
    return redirect("list_files")


@login_required
def download_file(request, file_id):
    try:
        blob = bucket.blob(file_id)
        if blob.exists():
            content = blob.download_as_bytes()
            response = HttpResponse(content, content_type=blob.content_type or "application/octet-stream")
            response['Content-Disposition'] = f'attachment; filename="{blob.name.split("/")[-1]}"'
            return response
    except Exception as e:
        return HttpResponse(f"Error al descargar archivo: {e}", status=500)
    return HttpResponse("Archivo no encontrado", status=404)


@login_required
def delete_file(request, file_id):
    try:
        blob = bucket.blob(file_id)
        if blob.exists():
            blob.delete()
    except Exception as e:
        return HttpResponse(f"Error al eliminar archivo: {e}", status=500)
    return redirect("list_files")


@login_required
def new_folder(request):
    if request.method == "POST":
        folder_name = request.POST.get("folder_name")
        current_folder = request.POST.get("current_folder", "")
        if folder_name:
            if not folder_name.endswith("/"):
                folder_name += "/"
            blob_name = f"{current_folder}{folder_name}" if current_folder else folder_name
            try:
                blob = bucket.blob(blob_name)
                blob.upload_from_string("")  # GCS no tiene carpetas reales
            except Exception as e:
                return HttpResponse(f"Error al crear carpeta: {e}", status=500)
    return redirect("list_files")




@login_required
def nuevo_requerimiento(request, proyecto_id):
    from plantillas_documentos_tecnicos.utils_documentos import inicializar_version_inicial
    from google.cloud import storage

    # ---------------------------------------------------------
    # 1) GET → Mostrar formulario nuevo_requerimiento_unico.html
    # ---------------------------------------------------------
    if request.method == "GET":

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    P.id, 
                    P.nombre AS proyecto,
                    CL.nombre AS cliente
                FROM proyectos P
                JOIN contratos C ON P.contrato_id = C.id
                JOIN clientes CL ON C.cliente_id = CL.id
                WHERE P.id = %s;
            """, [proyecto_id])
            row = cursor.fetchone()

        if not row:
            messages.error(request, "Proyecto no encontrado.")
            return redirect("usuario:lista_proyectos")

        proyecto_info = {
            "id": row[0],
            "nombre": row[1],
            "cliente": row[2],
        }

        # Cargar datos para el formulario
        with connection.cursor() as cursor:
            
            # Categorías
            cursor.execute("""
                SELECT id, nombre, descripcion 
                FROM categoria_documentos_tecnicos
                ORDER BY nombre
            """)
            grupos_maestros = [
                {"id": r[0], "nombre": r[1], "descripcion": r[2]}
                for r in cursor.fetchall()
            ]

            # Tipos de documento
            cursor.execute("""
                SELECT 
                    tdt.id,
                    tdt.categoria_id,
                    tdt.nombre,
                    CASE WHEN ptd.id IS NULL THEN FALSE ELSE TRUE END AS tiene_plantilla,
                    CASE WHEN ptd.version_actual_id IS NULL THEN FALSE ELSE TRUE END AS tiene_version
                FROM tipo_documentos_tecnicos tdt
                LEFT JOIN plantilla_tipo_doc ptd
                       ON ptd.tipo_documento_id = tdt.id
                ORDER BY tdt.nombre;
            """)
            documentos = [
                {
                    "id": r[0],
                    "categoria_id": r[1],
                    "nombre": r[2],
                    "tiene_plantilla": r[3],
                    "tiene_version": r[4],
                }
                for r in cursor.fetchall()
            ]

            # Usuarios
            cursor.execute("SELECT id, nombre, email FROM usuarios ORDER BY nombre")
            usuarios_todos = [
                {"id": r[0], "nombre": r[1], "email": r[2]}
                for r in cursor.fetchall()
            ]

        # Asignar documentos por categoría
        for g in grupos_maestros:
            g["documentos"] = [d for d in documentos if d["categoria_id"] == g["id"]]

        return render(request, "nuevo_requerimiento_unico.html", {
            "proyecto": proyecto_info,
            "grupos_maestros": grupos_maestros,
            "usuarios_todos": usuarios_todos,
        })

    # ---------------------------------------------------------
    # 2) POST → Crear requerimientos
    # ---------------------------------------------------------
    if request.method == "POST":

        usuario_id = request.user.id
        documentos_ids = request.POST.getlist("documentos_ids[]")

        if not documentos_ids:
            messages.error(request, "Debe seleccionar al menos un documento.")
            return redirect("usuario:detalle_proyecto", proyecto_id=proyecto_id)

        ROL_MAP = {"redactor": 1, "revisor": 2, "aprobador": 3}

        with connection.cursor() as cursor:

            # Obtener cliente + proyecto
            cursor.execute("""
                SELECT 
                    P.nombre,
                    CL.nombre
                FROM proyectos P
                JOIN contratos C ON P.contrato_id = C.id
                JOIN clientes CL ON C.cliente_id = CL.id
                WHERE P.id = %s
            """, [proyecto_id])
            proy_nom, cliente_nom = cursor.fetchone()

            cliente_nom = clean(cliente_nom)
            proy_nom = clean(proy_nom)

            # Cargar bucket
            storage_client = storage.Client()
            bucket = storage_client.bucket(settings.GCP_BUCKET_NAME)

            # Crear requerimientos uno por documento
            for doc_id in documentos_ids:

                restriccion = request.POST.get(f"restriccion_tipo_{doc_id}", "no_restringido")
                observaciones = request.POST.get(f"observaciones_{doc_id}", "").strip()

                h_ini = request.POST.get(f"fecha_inicio_elaboracion_{doc_id}")
                d_ini = request.POST.get(f"alertar_dias_inicio_{doc_id}")
                h_rev = request.POST.get(f"fecha_primera_revision_{doc_id}")
                d_rev = request.POST.get(f"alertar_dias_revision_{doc_id}")
                h_ent = request.POST.get(f"fecha_entrega_{doc_id}")
                d_ent = request.POST.get(f"alertar_dias_entrega_{doc_id}")

                # Crear requerimiento
                cursor.execute("""
                    INSERT INTO requerimiento_documento_tecnico
                    (proyecto_id, tipo_documento_id, fecha_registro, observaciones, confidencialidad)
                    VALUES (%s,%s,NOW(),%s,%s)
                    RETURNING id;
                """, [proyecto_id, doc_id, observaciones, restriccion])
                req_id = cursor.fetchone()[0]

                # Obtener categoría + tipo del documento
                cursor.execute("""
                    SELECT 
                        cdt.nombre AS categoria,
                        tdt.nombre AS tipo
                    FROM tipo_documentos_tecnicos tdt
                    JOIN categoria_documentos_tecnicos cdt
                        ON cdt.id = tdt.categoria_id
                    WHERE tdt.id = %s
                """, [doc_id])
                categoria_nom, tipo_nom = cursor.fetchone()

                categoria_nom = clean(categoria_nom)
                tipo_nom = clean(tipo_nom)

                # Construir rutas correctas
                base_path         = f"DocumentosProyectos/{cliente_nom}/{proy_nom}/"
                categoria_path    = f"{base_path}Documentos_Tecnicos/{categoria_nom}/"
                tipo_path         = f"{categoria_path}{tipo_nom}/"
                rq_path           = f"{tipo_path}RQ-{req_id}/"
                carpeta_plantilla = f"{rq_path}Plantilla/"

                # Crear carpetas
                bucket.blob(categoria_path).upload_from_string(b"")
                bucket.blob(tipo_path).upload_from_string(b"")
                bucket.blob(rq_path).upload_from_string(b"")
                bucket.blob(carpeta_plantilla).upload_from_string(b"")

                # Generar código del documento
                generar_codigo_documento(cursor, req_id)
                cursor.execute("""
                    SELECT codigo_documento
                    FROM requerimiento_documento_tecnico
                    WHERE id=%s
                """, [req_id])
                codigo = cursor.fetchone()[0]

                # Crear versión inicial
                inicializar_version_inicial(
                    cursor=cursor,
                    bucket=bucket,
                    requerimiento_id=req_id,
                    ruta_plantilla=carpeta_plantilla,
                    codigo_documento=codigo
                )

                # HITOS
                cursor.execute("""
                    INSERT INTO hitos_requerimiento_documento
                    (requerimiento_id,
                    fecha_inicio_elaboracion, alertar_dias_inicio,
                    fecha_primera_revision, alertar_dias_revision,
                    fecha_entrega, alertar_dias_entrega)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """, [req_id, h_ini, d_ini, h_rev, d_rev, h_ent, d_ent])

                # ROLES
                for rol_nombre in ["redactor", "revisor", "aprobador"]:
                    rol_id = ROL_MAP[rol_nombre]
                    for uid in request.POST.getlist(f"{rol_nombre}_id_{doc_id}[]"):
                        cursor.execute("""
                            INSERT INTO requerimiento_equipo_rol
                            (requerimiento_id, usuario_id, rol_id)
                            VALUES (%s,%s,%s)
                        """, [req_id, uid, rol_id])

                # LOG inicial
                cursor.execute("""
                    INSERT INTO log_estado_requerimiento_documento
                    (requerimiento_id, usuario_id, estado_origen_id, estado_destino_id,
                    fecha_cambio, observaciones)
                    VALUES (
                        %s, %s,
                        NULL,
                        (SELECT id FROM estado_documento WHERE nombre ILIKE 'Pendiente de Inicio' LIMIT 1),
                        NOW(),
                        'Plantilla inicial habilitada automáticamente.'
                    );
                """, [req_id, usuario_id])

        messages.success(request, "Requerimiento técnico creado correctamente.")
        return redirect("usuario:detalle_proyecto", proyecto_id=proyecto_id)










@login_required
def editar_requerimiento(request, requerimiento_id):

    # --- Obtener datos del requerimiento ---
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT rdt.id,
                   rdt.proyecto_id,
                   rdt.tipo_documento_id,
                   TDT.nombre,
                   rdt.observaciones,
                   rdt.confidencialidad
            FROM requerimiento_documento_tecnico rdt
            LEFT JOIN tipo_documentos_tecnicos TDT ON rdt.tipo_documento_id = TDT.id
            WHERE rdt.id = %s
        """, [requerimiento_id])
        row = cursor.fetchone()

    if not row:
        return render(request, "error.html", {"mensaje": "Requerimiento no encontrado."})

    requerimiento = {
        "id": row[0],
        "proyecto_id": row[1],
        "tipo_documento_id": row[2],
        "nombre_tipo_documento": row[3],
        "observaciones": row[4],
        "confidencialidad": row[5] or "no_restringido"
    }

    # --- Obtener usuarios y roles asignados ---
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, nombre FROM usuarios ORDER BY nombre;")
        usuarios = cursor.fetchall()

        cursor.execute("""
            SELECT usuario_id, rol_id
            FROM requerimiento_equipo_rol
            WHERE requerimiento_id = %s AND activo = TRUE
        """, [requerimiento_id])
        roles_asignados = cursor.fetchall()

    redactores = [u for u, r in roles_asignados if r == 1]
    revisores = [u for u, r in roles_asignados if r == 2]
    aprobadores = [u for u, r in roles_asignados if r == 3]

    # --- PROCESAR FORMULARIO ---
    if request.method == "POST":

        observaciones = request.POST.get("observaciones", "")
        redactores_post = request.POST.getlist("redactores")
        revisores_post = request.POST.getlist("revisores")
        aprobadores_post = request.POST.getlist("aprobadores")

        #  Nuevo: captura de confidencialidad
        confidencialidad = request.POST.get("confidencialidad", "no_restringido")

        # Seguridad — evita valores inválidos
        if confidencialidad not in ["confidencial", "restringido", "no_restringido"]:
            confidencialidad = "no_restringido"

        with connection.cursor() as cursor:

            # --- Actualizar requerimiento ---
            cursor.execute("""
                UPDATE requerimiento_documento_tecnico
                SET observaciones=%s,
                    confidencialidad=%s
                WHERE id=%s
            """, [observaciones, confidencialidad, requerimiento_id])

            # --- Desactivar todos los roles previos ---
            cursor.execute("""
                UPDATE requerimiento_equipo_rol
                SET activo=FALSE
                WHERE requerimiento_id=%s
            """, [requerimiento_id])

            # --- Insertar nuevos roles ---
            for u in redactores_post:
                cursor.execute("""
                    INSERT INTO requerimiento_equipo_rol
                    (requerimiento_id, usuario_id, rol_id, fecha_asignacion, activo)
                    VALUES (%s, %s, 1, NOW(), TRUE)
                """, [requerimiento_id, u])

            for u in revisores_post:
                cursor.execute("""
                    INSERT INTO requerimiento_equipo_rol
                    (requerimiento_id, usuario_id, rol_id, fecha_asignacion, activo)
                    VALUES (%s, %s, 2, NOW(), TRUE)
                """, [requerimiento_id, u])

            for u in aprobadores_post:
                cursor.execute("""
                    INSERT INTO requerimiento_equipo_rol
                    (requerimiento_id, usuario_id, rol_id, fecha_asignacion, activo)
                    VALUES (%s, %s, 3, NOW(), TRUE)
                """, [requerimiento_id, u])

        return redirect('usuario:detalle_proyecto', proyecto_id=requerimiento["proyecto_id"])

    return render(request, "editar_requerimiento.html", {
        "requerimiento": requerimiento,
        "usuarios": usuarios,
        "redactores": redactores,
        "revisores": revisores,
        "aprobadores": aprobadores
    })


@login_required
def eliminar_requerimiento(request, requerimiento_id):
    """
    Elimina un requerimiento técnico completo:
    - Archivos individuales
    - Carpeta RQ-XX completa
    - Carpeta del tipo si ya no quedan otros RQ del mismo tipo
    - Registros BD
    """

    try:
        # ============================================================
        # 1️⃣ Datos del requerimiento → usando NOMBRES reales (NO abreviaturas)
        # ============================================================
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    P.id AS proyecto_id,
                    P.nombre AS proyecto_nom_real,
                    CL.nombre AS cliente_nom_real,
                    CDT.nombre AS categoria_nom,
                    TDT.nombre AS tipo_nom,
                    R.tipo_documento_id
                FROM requerimiento_documento_tecnico R
                JOIN proyectos P ON R.proyecto_id = P.id
                JOIN contratos C ON P.contrato_id = C.id
                JOIN clientes CL ON C.cliente_id = CL.id
                JOIN tipo_documentos_tecnicos TDT ON R.tipo_documento_id = TDT.id
                JOIN categoria_documentos_tecnicos CDT ON TDT.categoria_id = CDT.id
                WHERE R.id = %s
            """, [requerimiento_id])

            row = cursor.fetchone()

        if not row:
            messages.error(request, "❌ Requerimiento no encontrado.")
            return redirect("usuario:lista_proyectos")

        (proyecto_id, proyecto_nom_real, cliente_nom_real,
         categoria_nom, tipo_nom, tipo_documento_id) = row

        # ============================================================
        # NORMALIZAR NOMBRES tal como al CREAR
        # ============================================================
        cliente_nom = clean(cliente_nom_real)
        proy_nom = clean(proyecto_nom_real)
        categoria_nom = clean(categoria_nom)
        tipo_nom = clean(tipo_nom)

        # ============================================================
        # PATHS correctos en GCS (coherentes con crear_proyecto)
        # ============================================================
        base_folder     = f"DocumentosProyectos/{cliente_nom}/{proy_nom}/"
        folder_tipo     = f"{base_folder}Documentos_Tecnicos/{categoria_nom}/{tipo_nom}/"
        folder_rq       = f"{folder_tipo}RQ-{requerimiento_id}/"

        storage_client = storage.Client()
        bucket = storage_client.bucket(settings.GCP_BUCKET_NAME)

        # ============================================================
        # 2️⃣ Eliminar archivos individuales via signed_url
        # ============================================================
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT signed_url
                FROM version_documento_tecnico
                WHERE requerimiento_documento_id = %s
            """, [requerimiento_id])

            signed_urls = [r[0] for r in cursor.fetchall() if r[0]]

        for url in signed_urls:
            blob_name = extract_blob_name_from_signed_url(url)
            if blob_name:
                blob = bucket.blob(blob_name)
                if blob.exists():
                    blob.delete()

        # ============================================================
        # 3️⃣ Eliminar carpeta RQ-x completa
        # ============================================================
        for blob in bucket.list_blobs(prefix=folder_rq):
            blob.delete()

        # ============================================================
        # 4️⃣ Verificar si quedan otros requerimientos del MISMO TIPO
        # ============================================================
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*)
                FROM requerimiento_documento_tecnico
                WHERE proyecto_id = %s
                  AND tipo_documento_id = %s
                  AND id <> %s
            """, [proyecto_id, tipo_documento_id, requerimiento_id])

            otros = cursor.fetchone()[0]

        # ============================================================
        # 5️⃣ Si no quedan → borrar el tipo completo
        # ============================================================
        if otros == 0:
            for blob in bucket.list_blobs(prefix=folder_tipo):
                blob.delete()

        # ============================================================
        # 6️⃣ Borrar documentos_generados asociados
        # ============================================================
        with connection.cursor() as cursor:
            cursor.execute("""
                DELETE FROM documentos_generados
                WHERE proyecto_id = %s
                  AND tipo_documento_id = %s
                  AND ruta_gcs LIKE %s
            """, [
                proyecto_id,
                tipo_documento_id,
                f"%/RQ-{requerimiento_id}/%"
            ])

        # ============================================================
        # 7️⃣ Eliminar requerimiento
        # ============================================================
        with connection.cursor() as cursor:
            cursor.execute("""
                DELETE FROM requerimiento_documento_tecnico 
                WHERE id = %s
            """, [requerimiento_id])

        messages.success(request, "🗑️ Requerimiento eliminado correctamente.")
        return redirect("usuario:detalle_proyecto", proyecto_id=proyecto_id)

    except Exception as e:
        print("⚠️ ERROR eliminar_requerimiento:", e)
        messages.error(request, f"Error al eliminar requerimiento: {e}")
        return redirect("usuario:lista_proyectos")







from django.http import JsonResponse
import json





from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import connection
# ================================
#  CREAR PROYECTO (VERSIÓN FINAL)
# ================================
@login_required
def crear_proyecto(request):
    from plantillas_documentos_tecnicos.utils_documentos import (
        inicializar_version_inicial,
        extract_blob_name_from_gcs_path
    )
    from google.cloud import storage

    if "proyecto_temp" not in request.session:
        request.session["proyecto_temp"] = {}

    temp = request.session["proyecto_temp"]
    paso_actual = request.POST.get("paso_actual", "1")
    accion = request.POST.get("accion")

    # =======================================================
    # Cargar catálogos y datos maestros
    # =======================================================
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, nombre, email 
            FROM usuarios
            WHERE cargo_id = 4 ORDER BY nombre
        """)
        usuarios_administrador = [
            {"id": r[0], "nombre": r[1], "email": r[2]} for r in cursor.fetchall()
        ]

        cursor.execute("SELECT id, nombre, email FROM usuarios ORDER BY nombre")
        usuarios_todos = [
            {"id": r[0], "nombre": r[1], "email": r[2]} for r in cursor.fetchall()
        ]

        cursor.execute("SELECT id, nombre, abreviatura FROM maquinas ORDER BY nombre")
        maquinas = [{"id": r[0], "nombre": r[1], "abreviatura": r[2] or ""} for r in cursor.fetchall()]

        cursor.execute("SELECT id, nombre FROM clientes ORDER BY nombre")
        clientes = [{"id": r[0], "nombre": r[1]} for r in cursor.fetchall()]

        cursor.execute("""
            SELECT id, cliente_id, nombre, ubicacion 
            FROM faenas ORDER BY nombre
        """)
        faenas = [{"id": r[0], "cliente_id": r[1], "nombre": r[2], "ubicacion": r[3]} for r in cursor.fetchall()]

        cursor.execute("""
            SELECT id, nombre, descripcion
            FROM categoria_documentos_tecnicos ORDER BY nombre
        """)
        grupos_maestros = [
            {"id": r[0], "nombre": r[1], "descripcion": r[2]}
            for r in cursor.fetchall()
        ]

        cursor.execute("""
            SELECT 
                tdt.id,
                tdt.categoria_id,
                tdt.nombre,
                CASE WHEN ptd.id IS NULL THEN FALSE ELSE TRUE END AS tiene_plantilla,
                CASE WHEN ptd.version_actual_id IS NULL THEN FALSE ELSE TRUE END AS tiene_version
            FROM tipo_documentos_tecnicos tdt
            LEFT JOIN plantilla_tipo_doc ptd
                   ON ptd.tipo_documento_id = tdt.id
            ORDER BY tdt.nombre;
        """)
        documentos = [
            {
                "id": r[0],
                "categoria_id": r[1],
                "nombre": r[2],
                "tiene_plantilla": r[3],
                "tiene_version": r[4],
            }
            for r in cursor.fetchall()
        ]

        cursor.execute("""
            SELECT c.id, c.numero_contrato, c.monto_total, c.fecha_firma,
                   c.representante_cliente_nombre, c.representante_cliente_correo,
                   c.representante_cliente_telefono, cl.id, cl.nombre
            FROM contratos c
            JOIN clientes cl ON cl.id = c.cliente_id
            ORDER BY c.numero_contrato
        """)
        contratos = [{
            "id": r[0], "numero_contrato": r[1], "monto_total": r[2],
            "fecha_firma": r[3], "representante_cliente_nombre": r[4],
            "representante_cliente_correo": r[5], "representante_cliente_telefono": r[6],
            "cliente_id": r[7], "cliente_nombre": r[8]
        } for r in cursor.fetchall()]

    # Relaciona documentos por categoría
    for g in grupos_maestros:
        g["documentos"] = [d for d in documentos if d["categoria_id"] == g["id"]]

    # =======================================================
    #   CONTROL DE PASOS
    # =======================================================
    if request.method == "POST":
        if accion == "anterior":
            paso_actual = str(max(1, int(paso_actual) - 1))

        elif accion == "siguiente":

            # PASO 1
            if paso_actual == "1":
                temp["paso1"] = {
                    "nombre": request.POST.get("nombre"),
                    "descripcion": request.POST.get("descripcion"),
                    "abreviatura": request.POST.get("abreviatura"),
                    "fecha_recepcion_evaluacion": request.POST.get("fecha_recepcion_evaluacion"),
                    "fecha_inicio_planificacion": request.POST.get("fecha_inicio_planificacion"),
                    "fecha_inicio_ejecucion": request.POST.get("fecha_inicio_ejecucion"),
                    "fecha_cierre_proyecto": request.POST.get("fecha_cierre_proyecto"),
                    "administrador": request.POST.get("administrador"),
                    "numero_servicio": request.POST.get("numero_servicio"),
                    "maquinas_ids": request.POST.getlist("maquinas_ids[]"),
                }

            # PASO 2
            elif paso_actual == "2":
                temp["paso2"] = {
                    "contrato_id": request.POST.get("contrato_id"),
                    "numero_contrato": request.POST.get("numero_contrato"),
                    "monto_total": request.POST.get("monto_total"),
                    "contrato_fecha_firma": request.POST.get("contrato_fecha_firma"),
                    "representante_cliente_nombre": request.POST.get("representante_cliente_nombre"),
                    "representante_cliente_correo": request.POST.get("representante_cliente_correo"),
                    "representante_cliente_telefono": request.POST.get("representante_cliente_telefono"),
                    "cliente_id": request.POST.get("cliente_id"),
                    "cliente_nombre": request.POST.get("cliente_nombre"),
                    "cliente_abreviatura": request.POST.get("cliente_abreviatura"),
                    "cliente_rut": request.POST.get("cliente_rut"),
                    "cliente_direccion": request.POST.get("cliente_direccion"),
                    "cliente_correo": request.POST.get("cliente_correo"),
                    "cliente_telefono": request.POST.get("cliente_telefono"),
                    "faena_id": request.POST.get("faena_id"),
                    "faena_nombre": request.POST.get("faena_nombre"),
                    "faena_ubicacion": request.POST.get("faena_ubicacion"),
                }

            # PASO 3
            elif paso_actual == "3":

                documentos_roles = {}
                for doc_id in request.POST.getlist("documentos_ids[]"):
                    documentos_roles[doc_id] = {
                        "redactores": request.POST.getlist(f"redactor_id_{doc_id}[]"),
                        "revisores": request.POST.getlist(f"revisor_id_{doc_id}[]"),
                        "aprobadores": request.POST.getlist(f"aprobador_id_{doc_id}[]"),
                        "observaciones": request.POST.get(f"observaciones_{doc_id}", ""),
                        "restriccion": request.POST.get(f"restriccion_tipo_{doc_id}", "no_restringido"),
                        "hitos": {
                            "fecha_inicio_elaboracion": request.POST.get(f"fecha_inicio_elaboracion_{doc_id}"),
                            "alertar_dias_inicio": request.POST.get(f"alertar_dias_inicio_{doc_id}"),
                            "fecha_primera_revision": request.POST.get(f"fecha_primera_revision_{doc_id}"),
                            "alertar_dias_revision": request.POST.get(f"alertar_dias_revision_{doc_id}"),
                            "fecha_entrega": request.POST.get(f"fecha_entrega_{doc_id}"),
                            "alertar_dias_entrega": request.POST.get(f"alertar_dias_entrega_{doc_id}"),
                        }
                    }

                temp["paso3"] = {
                    "observaciones": request.POST.get("observaciones"),
                    "documentos_ids": request.POST.getlist("documentos_ids[]"),
                    "documentos_roles": documentos_roles
                }

            request.session.modified = True
            paso_actual = str(min(4, int(paso_actual) + 1))

        # =======================================================
        #   CONFIRMAR — PASO 4
        # =======================================================
        elif accion == "confirmar":

            resumen = {
                **temp.get("paso1", {}),
                **temp.get("paso2", {}),
                **temp.get("paso3", {})
            }

            # --------------------- STORAGE ---------------------
            storage_client = storage.Client()
            bucket = storage_client.bucket(settings.GCP_BUCKET_NAME)

            cliente_nom = clean(resumen.get("cliente_nombre") or "Cliente")
            proyecto_nom = clean(resumen.get("nombre") or "Proyecto")

            # Carpeta raíz del proyecto
            base_path = f"DocumentosProyectos/{cliente_nom}/{proyecto_nom}/"
            bucket.blob(base_path).upload_from_string(b"")

            documentos_roles = resumen.get("documentos_roles", {})

            try:
                with transaction.atomic():
                    with connection.cursor() as cursor:

                        # ========================================
                        # CLIENTE
                        # ========================================
                        cliente_id = resumen.get("cliente_id")
                        if not cliente_id:
                            cursor.execute("""
                                INSERT INTO clientes
                                (nombre, abreviatura, rut, direccion, correo_contacto, telefono_contacto)
                                VALUES (%s,%s,%s,%s,%s,%s)
                                RETURNING id;
                            """, [
                                resumen.get("cliente_nombre"),
                                resumen.get("cliente_abreviatura"),
                                resumen.get("cliente_rut"),
                                resumen.get("cliente_direccion"),
                                resumen.get("cliente_correo"),
                                resumen.get("cliente_telefono"),
                            ])
                            cliente_id = cursor.fetchone()[0]

                        # ========================================
                        # FAENA
                        # ========================================
                        faena_id = resumen.get("faena_id")
                        if not faena_id:
                            cursor.execute("""
                                INSERT INTO faenas (cliente_id, nombre, ubicacion)
                                VALUES (%s,%s,%s)
                                RETURNING id;
                            """, [
                                cliente_id,
                                resumen.get("faena_nombre"),
                                resumen.get("faena_ubicacion"),
                            ])
                            faena_id = cursor.fetchone()[0]

                        # ========================================
                        # CONTRATO
                        # ========================================
                        contrato_id = resumen.get("contrato_id")
                        if not contrato_id:
                            cursor.execute("""
                                INSERT INTO contratos(
                                    numero_contrato, monto_total, fecha_firma,
                                    representante_cliente_nombre, representante_cliente_correo,
                                    representante_cliente_telefono, cliente_id
                                )
                                VALUES (%s,%s,%s,%s,%s,%s,%s)
                                RETURNING id;
                            """, [
                                resumen.get("numero_contrato"),
                                resumen.get("monto_total"),
                                resumen.get("contrato_fecha_firma"),
                                resumen.get("representante_cliente_nombre"),
                                resumen.get("representante_cliente_correo"),
                                resumen.get("representante_cliente_telefono"),
                                cliente_id
                            ])
                            contrato_id = cursor.fetchone()[0]

                        # ========================================
                        # PROYECTO
                        # ========================================
                        cursor.execute("""
                            INSERT INTO proyectos(
                                nombre, descripcion, abreviatura, numero_servicio,
                                contrato_id, faena_id, administrador_id,
                                fecha_recepcion_evaluacion, fecha_inicio_planificacion,
                                fecha_inicio_ejecucion, fecha_cierre_proyecto, path_gcs
                            )
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            RETURNING id;
                        """, [
                            resumen.get("nombre"),
                            resumen.get("descripcion"),
                            resumen.get("abreviatura"),
                            resumen.get("numero_servicio"),
                            contrato_id,
                            faena_id,
                            resumen.get("administrador"),
                            resumen.get("fecha_recepcion_evaluacion"),
                            resumen.get("fecha_inicio_planificacion"),
                            resumen.get("fecha_inicio_ejecucion"),
                            resumen.get("fecha_cierre_proyecto"),
                            base_path,
                        ])
                        proyecto_id = cursor.fetchone()[0]

                        # ========================================
                        # MÁQUINAS
                        # ========================================
                        for m_id in temp.get("paso1", {}).get("maquinas_ids", []):
                            cursor.execute("""
                                INSERT INTO proyecto_maquina (proyecto_id, maquina_id)
                                VALUES (%s, %s)
                            """, [proyecto_id, m_id])

                        # ==========================================================
                        #  REQUERIMIENTOS + ÁRBOL COMPLETO DE GCS + VERSIÓN INICIAL
                        # ==========================================================
                        for doc_id, datos in documentos_roles.items():

                            # ----------------------------------------------
                            # Crear RQ en DB
                            # ----------------------------------------------
                            cursor.execute("""
                                INSERT INTO requerimiento_documento_tecnico(
                                    proyecto_id, tipo_documento_id, fecha_registro,
                                    observaciones, confidencialidad
                                )
                                VALUES (%s,%s,NOW(),%s,%s)
                                RETURNING id;
                            """, [
                                proyecto_id,
                                doc_id,
                                datos.get("observaciones"),
                                datos.get("restriccion", "no_restringido"),
                            ])
                            req_id = cursor.fetchone()[0]

                            # ----------------------------------------------
                            # Código del documento
                            # ----------------------------------------------
                            generar_codigo_documento(cursor, req_id)
                            cursor.execute(
                                "SELECT codigo_documento FROM requerimiento_documento_tecnico WHERE id=%s",
                                [req_id]
                            )
                            codigo_documento = cursor.fetchone()[0]

                            # =======================================================
                            #        ÁRBOL DE CARPETAS CORRECTO
                            # =======================================================

                            # Obtener categoría y tipo
                            cursor.execute("""
                                SELECT 
                                    cdt.nombre AS categoria,
                                    tdt.nombre AS tipo
                                FROM tipo_documentos_tecnicos tdt
                                JOIN categoria_documentos_tecnicos cdt
                                     ON cdt.id = tdt.categoria_id
                                WHERE tdt.id = %s
                            """, [doc_id])
                            categoria_nom, tipo_nom = cursor.fetchone()

                            categoria_nom = clean(categoria_nom)
                            tipo_nom = clean(tipo_nom)

                            categoria_path = f"{base_path}Documentos_Tecnicos/{categoria_nom}/"
                            tipo_path      = f"{categoria_path}{tipo_nom}/"
                            rq_path        = f"{tipo_path}RQ-{req_id}/"
                            plantilla_path = f"{rq_path}Plantilla/"

                            # Crear carpetas en cascada
                            bucket.blob(categoria_path).upload_from_string(b"")
                            bucket.blob(tipo_path).upload_from_string(b"")
                            bucket.blob(rq_path).upload_from_string(b"")
                            bucket.blob(plantilla_path).upload_from_string(b"")

                            # =======================================================
                            #   Crear VERSIÓN INICIAL (100% compatible con tu sistema)
                            # =======================================================
                            inicializar_version_inicial(
                                cursor=cursor,
                                bucket=bucket,
                                requerimiento_id=req_id,
                                ruta_plantilla=plantilla_path,
                                codigo_documento=codigo_documento
                            )

                            # ----------------------------------------------
                            # HITOS
                            # ----------------------------------------------
                            h = datos.get("hitos", {})
                            cursor.execute("""
                                INSERT INTO hitos_requerimiento_documento
                                (requerimiento_id,
                                 fecha_inicio_elaboracion, alertar_dias_inicio,
                                 fecha_primera_revision, alertar_dias_revision,
                                 fecha_entrega, alertar_dias_entrega)
                                VALUES (%s,%s,%s,%s,%s,%s,%s);
                            """, [
                                req_id,
                                h.get("fecha_inicio_elaboracion"),
                                h.get("alertar_dias_inicio"),
                                h.get("fecha_primera_revision"),
                                h.get("alertar_dias_revision"),
                                h.get("fecha_entrega"),
                                h.get("alertar_dias_entrega"),
                            ])

                            # ----------------------------------------------
                            # ROLES
                            # ----------------------------------------------
                            ROL_MAP = {"redactor": 1, "revisor": 2, "aprobador": 3}

                            for rol_nombre, usuarios_ids in [
                                ("redactor", datos.get("redactores", [])),
                                ("revisor", datos.get("revisores", [])),
                                ("aprobador", datos.get("aprobadores", [])),
                            ]:
                                rol_id = ROL_MAP[rol_nombre]
                                for uid in usuarios_ids:
                                    cursor.execute("""
                                        INSERT INTO requerimiento_equipo_rol
                                        (requerimiento_id, usuario_id, rol_id)
                                        VALUES (%s, %s, %s)
                                    """, [req_id, uid, rol_id])

                            # ----------------------------------------------
                            # LOG INICIAL
                            # ----------------------------------------------
                            cursor.execute("""
                                INSERT INTO log_estado_requerimiento_documento
                                (requerimiento_id, usuario_id, estado_origen_id, estado_destino_id,
                                 fecha_cambio, observaciones)
                                VALUES (
                                    %s, %s, NULL,
                                    (SELECT id FROM estado_documento WHERE nombre ILIKE 'Pendiente de Inicio' LIMIT 1),
                                    NOW(),
                                    'Plantilla inicial habilitada automáticamente.'
                                );
                            """, [req_id, request.user.id])

                # ----------------------------------------------
                # ÉXITO
                # ----------------------------------------------
                del request.session["proyecto_temp"]
                messages.success(request, "Proyecto creado correctamente.")
                return redirect("usuario:lista_proyectos")

            except Exception as e:
                traceback.print_exc()
                messages.error(request, f"Error al crear proyecto: {e}")

    # =============================
    #  RENDER PASOS / CONFIRMACIÓN
    # =============================
    steps = ["Datos Generales", "Contrato y Cliente", "Responsables y Documentos", "Confirmación"]
    resumen = temp.get("paso1", {})

    if paso_actual == "4":
        resumen = {
            **temp.get("paso1", {}),
            **temp.get("paso2", {}),
            **temp.get("paso3", {}),
        }

        documentos_roles = resumen.get("documentos_roles", {})

        with connection.cursor() as cursor:
            completo = {}
            for doc_id, datos in documentos_roles.items():
                cursor.execute("SELECT nombre FROM tipo_documentos_tecnicos WHERE id=%s", [doc_id])
                row = cursor.fetchone()
                completo[doc_id] = {
                    "documento_nombre": row[0] if row else f"Documento {doc_id}",
                    **datos
                }
            resumen["documentos_roles"] = completo

    return render(request, "crear_proyecto.html", {
        "paso_actual": paso_actual,
        "proyecto_temp": temp.get("paso1", {}),
        "steps": steps,
        "resumen": resumen,
        "usuarios_administrador": usuarios_administrador,
        "usuarios_todos": usuarios_todos,
        "maquinas": maquinas,
        "contratos": contratos,
        "clientes": clientes,
        "faenas": faenas,
        "grupos_maestros": grupos_maestros,
        "documentos": documentos,
    })



####################################Crear Proyecto - Paso 2 ##################################

@login_required
def obtener_datos_contrato(request, contrato_id):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT c.id, c.numero_contrato, c.monto_total, c.fecha_firma,
                   c.representante_cliente_nombre, c.representante_cliente_correo,
                   c.representante_cliente_telefono,
                   cl.id AS cliente_id, cl.nombre AS cliente_nombre,
                   cl.abreviatura, cl.rut, cl.direccion, cl.correo_contacto, cl.telefono_contacto
            FROM contratos c
            JOIN clientes cl ON cl.id = c.cliente_id
            WHERE c.id = %s
        """, [contrato_id])
        r = cursor.fetchone()
    if not r:
        return JsonResponse({}, status=404)
    return JsonResponse({
        "contrato_id": r[0],
        "numero_contrato": r[1],
        "monto_total": r[2],
        "fecha_firma": r[3],
        "representante_cliente_nombre": r[4],
        "representante_cliente_correo": r[5],
        "representante_cliente_telefono": r[6],
        "cliente_id": r[7],
        "cliente_nombre": r[8],
        "cliente_abreviatura": r[9],
        "cliente_rut": r[10],
        "cliente_direccion": r[11],
        "cliente_correo": r[12],
        "cliente_telefono": r[13],
    })

@login_required
def obtener_datos_cliente(request, cliente_id):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, nombre, abreviatura, rut, direccion, correo_contacto, telefono_contacto
            FROM clientes WHERE id = %s
        """, [cliente_id])
        r = cursor.fetchone()
    if not r:
        return JsonResponse({}, status=404)
    return JsonResponse({
        "cliente_id": r[0],
        "nombre": r[1],
        "abreviatura": r[2],
        "rut": r[3],
        "direccion": r[4],
        "correo": r[5],
        "telefono": r[6],
    })

@login_required
def obtener_datos_faena(request, faena_id):
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, nombre, ubicacion FROM faenas WHERE id = %s", [faena_id])
        r = cursor.fetchone()
    if not r:
        return JsonResponse({}, status=404)
    return JsonResponse({
        "faena_id": r[0],
        "nombre": r[1],
        "ubicacion": r[2],
    })





######################################################################




@csrf_exempt
@login_required
def generar_abreviatura_proyecto(request):
    """
    Genera la abreviatura de un proyecto a partir de:
    - nombre de la máquina
    - descripción
    - fecha de recepción / evaluación
    """
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        import json, datetime, re
        data = json.loads(request.body)

        maquina = data.get("maquina", "").strip()
        descripcion = data.get("descripcion", "").strip().upper()
        fecha = data.get("fecha_recepcion_evaluacion", "").strip()

        print(" Datos recibidos:", data)  # DEBUG

        if not maquina or not fecha:
            return JsonResponse({"abreviatura": ""})

        # Limpiar nombre de máquina (sin abreviatura entre paréntesis)
        maquina = re.sub(r"\(.*?\)", "", maquina).strip().upper().replace(" ", "")

        # Limpiar descripción (usar primeras 2 palabras)
        palabras = [p for p in descripcion.split() if len(p) > 2]
        descripcion_limpia = "".join(palabras[:2]).upper()

        # Formato fecha MMYY
        try:
            fecha_obj = datetime.datetime.strptime(fecha, "%Y-%m-%d")
            mes = f"{fecha_obj.month:02}"
            anio = str(fecha_obj.year)[-2:]
            fecha_formato = f"{mes}{anio}"
        except ValueError:
            fecha_formato = ""

        abreviatura = f"{maquina}.{descripcion_limpia}.{fecha_formato}"
        print("✅ Abreviatura generada:", abreviatura)

        return JsonResponse({"abreviatura": abreviatura})

    except Exception as e:
        print("❌ Error generar_abreviatura_proyecto:", e)
        return JsonResponse({"error": str(e)}, status=400)
#paso2 ####################  ####################    ####################    ####################  


@csrf_exempt
@login_required
def generar_abreviatura_cliente(request):
    """
    Genera la abreviatura de un cliente (vía AJAX) a partir de su nombre.
    Equivalente al modelo del proyecto: una sola función con request.
    """
    if request.method != "GET":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        nombre = request.GET.get("nombre", "").strip()
        if not nombre:
            return JsonResponse({"abreviatura": ""})

        # --- Normalización y limpieza ---
        nombre = unidecode.unidecode(nombre.upper().strip())
        nombre = re.sub(r'[^A-ZÑ\s]', ' ', nombre)
        nombre = re.sub(r'\s+', ' ', nombre).strip()
        palabras = nombre.split()

        stopwords = {
            "DE", "DEL", "LA", "LOS", "LAS", "Y", "E", "S", "SA", "SAA",
            "LTDA", "LIMITADA", "COMPANIA", "COMPAÑIA", "CORPORACION",
            "CORPORACIÓN", "EMPRESA", "GRUPO", "INDUSTRIAL", "SERVICIOS"
        }
        palabras = [p for p in palabras if p not in stopwords]

        # --- Generación base ---
        if not palabras:
            base = "GEN"
        elif len(palabras) == 1:
            base = palabras[0][:6]
        else:
            partes = [p[:3] for p in palabras[:3]]
            base = ''.join(partes).upper()
            base = base[:6] if len(base) > 6 else base

        abrev_final = base
        contador = 1

        # --- Validación de unicidad ---
        with connection.cursor() as cursor:
            while True:
                cursor.execute("SELECT COUNT(*) FROM clientes WHERE abreviatura = %s;", [abrev_final])
                if cursor.fetchone()[0] == 0:
                    break
                abrev_final = f"{base}{contador}"
                contador += 1

        return JsonResponse({"abreviatura": abrev_final})

    except Exception as e:
        print("❌ Error generar_abreviatura_cliente:", e)
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@csrf_exempt
def obtener_abreviatura_cliente(request, cliente_id):
    """
    Devuelve la abreviatura registrada para un cliente existente.
    """
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT abreviatura FROM clientes WHERE id = %s;", [cliente_id])
            result = cursor.fetchone()
            if result and result[0]:
                return JsonResponse({"abreviatura": result[0]})
            else:
                return JsonResponse({"abreviatura": None})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    



def obtener_numordenservicio(path_archivo):
    wb = load_workbook(path_archivo, data_only=True)

    # Recorre todos los nombres definidos en el libro
    for name, defn in wb.defined_names.items():
        if name.lower() == "numordenservicio":
            try:
                for title, coord in defn.destinations:
                    sheet = wb[title]
                    cell = sheet[coord]
                    return str(cell.value or "")
            except Exception as e:
                print(f"⚠️ Error leyendo numordenservicio: {e}")
                return ""
    return ""

@login_required
def leer_excel_numero_servicio(request):
    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES["archivo"]
        import tempfile, os

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:

            for chunk in archivo.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        numero_servicio = obtener_numordenservicio(tmp_path)
        os.remove(tmp_path)
        return JsonResponse({"numero_servicio": numero_servicio})
    return JsonResponse({"numero_servicio": ""})


@login_required
def editar_usuario(request, user_id):
    # Traer usuario
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, nombre, email, email_secundario,
                   telefono_corporativo, telefono_secundario, area_trabajo, cargo_id
            FROM usuarios
            WHERE id = %s
        """, [user_id])
        row = cursor.fetchone()

    if not row:
        messages.error(request, "Usuario no encontrado.")
        return redirect('usuario:lista_usuarios')

    usuario = {
        'id': row[0],
        'nombre': row[1],
        'email': row[2],
        'email_secundario': row[3],
        'telefono_corporativo': row[4],
        'telefono_secundario': row[5],
        'area_trabajo': row[6],
        'cargo_id': row[7],
    }

    # Traer áreas
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, nombre FROM area_cargo_empresa")
        areas = cursor.fetchall()

    # Traer todos los cargos
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, nombre, area_id FROM cargos_empresa")
        cargos = cursor.fetchall()

    # Convertir a diccionarios
    areas = [{'id': a[0], 'nombre': a[1]} for a in areas]
    cargos = [{'id': c[0], 'nombre': c[1], 'area_id': c[2]} for c in cargos]

    if request.method == "POST":
        nombre = request.POST.get('nombre', usuario['nombre'])
        email = request.POST.get('email', usuario['email'])
        email_secundario = request.POST.get('email_secundario', usuario['email_secundario'])
        telefono_corporativo = request.POST.get('telefono_corporativo', usuario['telefono_corporativo'])
        telefono_secundario = request.POST.get('telefono_secundario', usuario['telefono_secundario'])
        area_trabajo_id = request.POST.get('area_trabajo')
        cargo_id = request.POST.get('cargo_id')

        # Guardar area_trabajo como nombre de área
        area_nombre = next((a['nombre'] for a in areas if str(a['id'])==area_trabajo_id), '')

        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE usuarios
                SET nombre=%s, email=%s, email_secundario=%s,
                    telefono_corporativo=%s, telefono_secundario=%s,
                    area_trabajo=%s, cargo_id=%s
                WHERE id=%s
            """, [nombre, email, email_secundario,
                  telefono_corporativo, telefono_secundario,
                  area_nombre, cargo_id, user_id])

        messages.success(request, f"Usuario {nombre} actualizado correctamente.")
        return redirect('usuario:lista_usuarios')

    return render(request, 'editar_usuario.html', {
        'usuario': usuario,
        'areas': areas,
        'cargos': cargos,
    })



@login_required
def ver_estadisticas_usuario(request, user_id):
    # Obtener datos del usuario
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, nombre, email, email_secundario,
                   telefono_corporativo, telefono_secundario,
                   area_trabajo, cargo_id, fecha_registro
            FROM usuarios
            WHERE id = %s
        """, [user_id])
        row = cursor.fetchone()

    if not row:
        messages.error(request, "Usuario no encontrado.")
        return redirect('usuario:lista_usuarios')

    usuario = {
        'id': row[0],
        'nombre': row[1],
        'email_corporativo': row[2],
        'email_secundario': row[3],
        'telefono_corporativo': row[4],
        'telefono_secundario': row[5],
        'area_trabajo': row[6],
        'cargo': row[7],  # Opcional: traducir cargo_id a nombre después
        'fecha_registro': row[8],
    }

    # Por ahora, solo renderizamos el HTML vacío de estadísticas
    return render(request, 'usuario_estadisticas.html', {
        'usuario': usuario,
    })


@login_required
def editar_proyecto(request, proyecto_id):
    # Obtener datos del proyecto incluyendo el id del administrador
    sql = """
    SELECT
        P.id AS proyecto_id,
        P.nombre AS nombre_proyecto,
        P.descripcion AS proyecto_descripcion,
        P.numero_servicio,
        P.fecha_recepcion_evaluacion,
        P.fecha_cierre_proyecto,
        U.id AS administrador_id,
        U.nombre AS administrador_nombre_completo,
        U.email AS administrador_email
    FROM public.proyectos P
    LEFT JOIN public.usuarios U ON P.administrador_id = U.id
    WHERE P.id = %s
    """
    
    with connection.cursor() as cursor:
        cursor.execute(sql, [proyecto_id])
        row = cursor.fetchone()
        if not row:
            raise Http404("Proyecto no encontrado")

        columns = [col[0] for col in cursor.description]
        proyecto = dict(zip(columns, row))

        # Lista de administradores con cargo_id = 4
        cursor.execute("""
            SELECT u.id, u.nombre, u.email
            FROM public.usuarios u
            WHERE u.cargo_id = 4
            ORDER BY u.nombre
        """)
        administradores = [
            dict(zip([col[0] for col in cursor.description], r))
            for r in cursor.fetchall()
        ]

    # Formatear fechas para input type="date"
    proyecto['fecha_recepcion_evaluacion'] = (
        proyecto['fecha_recepcion_evaluacion'].strftime('%Y-%m-%d')
        if proyecto['fecha_recepcion_evaluacion'] else ''
    )
    proyecto['fecha_cierre_proyecto'] = (
        proyecto['fecha_cierre_proyecto'].strftime('%Y-%m-%d')
        if proyecto['fecha_cierre_proyecto'] else ''
    )

    context = {
        'proyecto': proyecto,
        'administradores': administradores
    }
    return render(request, "editar_proyecto.html", context)



def editar_contrato(request, contrato_id):
    contrato = get_object_or_404(Contrato, id=contrato_id)
    return render(request, 'usuario/editar_contrato.html', {'contrato': contrato})

def editar_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    return render(request, 'usuario/editar_cliente.html', {'cliente': cliente})

@login_required
def editar_maquina(request, maquina_id):
    with connection.cursor() as cursor:
        # Buscar los datos de la máquina
        cursor.execute("""
            SELECT m.id, m.nombre, m.abreviatura, m.marca, m.modelo, m.anio_fabricacion,
                   m.tipo, m.descripcion, pm.proyecto_id
            FROM maquinas m
            LEFT JOIN proyecto_maquina pm ON pm.maquina_id = m.id
            WHERE m.id = %s
        """, [maquina_id])
        row = cursor.fetchone()

    if not row:
        raise Http404("Máquina no encontrada")

    maquina = {
        'id': row[0],
        'nombre': row[1],
        'abreviatura': row[2],
        'marca': row[3],
        'modelo': row[4],
        'anio_fabricacion': row[5],
        'tipo': row[6],
        'descripcion': row[7],
        'proyecto_id': row[8],  # ← tomado desde proyecto_maquina
    }

    return render(request, 'editar_maquina.html', {'maquina': maquina})


def generar_codigo_documento(cursor, req_id):
    """
    Genera código_documento usando abreviaturas, no nombres.
    Formato:
    CLIENTE_ABREV - PROYECTO_ABREV - CATEG_ABREV - TIPO_ABREV - RQ<ID>
    """

    cursor.execute("""
        SELECT 
            CL.abreviatura AS cliente_abrev,
            P.abreviatura AS proyecto_abrev,
            CDT.abreviatura AS categoria_abrev,
            TDT.abreviatura AS tipo_abrev
        FROM requerimiento_documento_tecnico R
        JOIN proyectos P ON R.proyecto_id = P.id
        JOIN contratos C ON P.contrato_id = C.id
        JOIN clientes CL ON C.cliente_id = CL.id
        JOIN tipo_documentos_tecnicos TDT ON R.tipo_documento_id = TDT.id
        JOIN categoria_documentos_tecnicos CDT ON TDT.categoria_id = CDT.id
        WHERE R.id = %s
    """, [req_id])

    row = cursor.fetchone()
    cliente_abrev, proyecto_abrev, categoria_abrev, tipo_abrev = row

    # Si alguna abreviatura viene NULL → reemplazar
    cliente_abrev = cliente_abrev or "CLT"
    proyecto_abrev = proyecto_abrev or "PRY"
    categoria_abrev = categoria_abrev or "CAT"
    tipo_abrev = tipo_abrev or "DOC"

    codigo = f"{cliente_abrev}-{proyecto_abrev}-{categoria_abrev}-{tipo_abrev}-RQ{req_id}"

    cursor.execute("""
        UPDATE requerimiento_documento_tecnico
        SET codigo_documento = %s
        WHERE id = %s
    """, [codigo, req_id])

