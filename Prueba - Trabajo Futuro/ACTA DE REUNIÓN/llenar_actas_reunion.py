# ===============================================================
# generar_actas_reunion_desde_json.py — VERSIÓN UNIVERSAL FINAL
# Usa el JSON generado por leer_universal.py para rellenar el DOCX.
# Mantiene: controles, excels embebidos, estilos, tablas.
# ===============================================================

import os
import json
import random
from datetime import datetime, timedelta
import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.table import Table

NS_W = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

# ---------------------------------------------------------------
# Catálogos
# ---------------------------------------------------------------
PROYECTOS = [
    "Overhaul Chancador Primario MP1000",
    "Mantención Harnero Banana DF501",
    "Reparación Molino SAG 40x22",
    "Reemplazo Sistema Hidráulico Chancador Secundario HP300",
    "Servicio Major Overhaul Chancador GP500"
]

ADMINISTRADORES = [
    "Carlos Muñoz", "Francisca Rivas", "Pedro González",
    "Ana Villalobos", "Luis Arrieta", "Javiera Troncoso"
]

CLIENTES = [
    "Codelco Andina", "BHP Escondida", "Collahuasi",
    "Anglo American Los Bronces", "Minera Centinela", "Minera Spence"
]


# ======================================================================
# COPIAR ESTILOS EXACTOS
# ======================================================================
def copy_cell_style(src, dst):
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


# ======================================================================
# INSERTAR FILA EN TABLA
# ======================================================================
def insertar_fila_tabla(ws, tabla, valores):
    min_col, min_row, max_col, max_row = range_boundaries(tabla.ref)
    insert_pos = max_row
    ws.insert_rows(insert_pos)
    max_row += 1
    tabla.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    for j, val in enumerate(valores):
        col = get_column_letter(min_col + j)
        cell = ws[f"{col}{insert_pos}"]

        # copiar estilo desde la fila 2 (min_row+1)
        style_src = ws[f"{col}{min_row + 1}"]
        copy_cell_style(style_src, cell)

        cell.value = val


# ======================================================================
# CREAR TABLA SI NO EXISTE (listobject)
# ======================================================================
def get_or_create_table(ws, tabla_info):
    for t in ws._tables.values():
        if t.displayName == tabla_info["tabla"]:
            return t

    nueva = Table(displayName=tabla_info["tabla"], ref=tabla_info["rango"])
    ws.add_table(nueva)
    return nueva


# ======================================================================
# RELLENAR EXCEL EMBEBIDO UTILIZANDO JSON UNIVERSAL
# ======================================================================
def rellenar_excel_embebido(xlsx_bytes, excel_desc):
    wb = load_workbook(filename=BytesIO(xlsx_bytes))

    for tinfo in excel_desc["tablas"]:
        ws = wb[tinfo["worksheet"]]
        tabla = get_or_create_table(ws, tinfo)
        columnas = tinfo["columnas"]

        min_col, min_row, max_col, max_row = range_boundaries(tabla.ref)

        # -----------------------------------------------------------------
        # RELLENAR PRIMERA FILA EXISTENTE (registros[0])
        # -----------------------------------------------------------------
        def generar_valores(k):
            if tinfo["tabla"] == "participantes_metso":
                return [f"Metso Persona {k}", random.choice(["Supervisor", "Ingeniero", "Técnico"])]

            elif tinfo["tabla"] == "participantes_cliente":
                return [f"Cliente Persona {k}", random.choice(["Jefe Turno", "Planificador"])]

            elif tinfo["tabla"] == "propositos_reunion":
                return [k, f"Propósito generado automáticamente {k}"]

            elif tinfo["tabla"] == "temas_tratados":
                fecha = datetime.today() + timedelta(days=k)
                return [
                    k,
                    f"Tema tratado {k}",
                    random.choice(["Juan Pérez", "Ana Soto", "Luis Rivas"]),
                    fecha.strftime("%Y-%m-%d"),
                    random.randint(0, 100) / 100
                ]

            return [None] * len(columnas)

        # FILA BASE
        base_vals = generar_valores(1)
        for j, v in enumerate(base_vals):
            col = get_column_letter(min_col + j)
            cell = ws[f"{col}{max_row}"]
            style_src = ws[f"{col}{min_row + 1}"]
            copy_cell_style(style_src, cell)
            cell.value = v

        # FILAS ADICIONALES
        for k in range(2, 6):
            insertar_fila_tabla(ws, tabla, generar_valores(k))

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ======================================================================
# GENERAR DATOS SINTÉTICOS
# ======================================================================
def generar_datos_sinteticos(indice_doc, total_docs):
    fecha_base = datetime.today() - timedelta(days=total_docs)
    fecha_doc = fecha_base + timedelta(days=indice_doc)

    return {
        "nombre_proyecto": random.choice(PROYECTOS),
        "fecha_realizacion_reunion": fecha_doc.strftime("%Y-%m-%d"),
        "administrador_servicio": random.choice(ADMINISTRADORES),
        "nombre_cliente": random.choice(CLIENTES),
        "num_servicio": f"SERV{random.randint(1000000,9999999)}",
        "lugar_reunion": random.choice(["Online", "Presencial"])
    }


# ======================================================================
# RELLENAR CONTROLES DEL DOCX USANDO JSON UNIVERSAL
# ======================================================================
def fill_content_controls(xml_bytes, context):
    root = ET.fromstring(xml_bytes)

    for sdt in root.findall(".//w:sdt", NS_W):

        tag_el = sdt.find("./w:sdtPr/w:tag", NS_W)
        if tag_el is None:
            continue

        key = tag_el.get("{%s}val" % NS_W["w"])
        if key not in context:
            continue

        nuevo = str(context[key])

        content = sdt.find("./w:sdtContent", NS_W)
        if content is None:
            continue

        # limpiar
        for t in content.findall(".//w:t", NS_W):
            t.text = ""

        run = content.find(".//w:r", NS_W)
        if run is None:
            run = ET.SubElement(content, "{%s}r" % NS_W["w"])

        t_final = run.find("w:t", NS_W)
        if t_final is None:
            t_final = ET.SubElement(run, "{%s}t" % NS_W["w"])

        t_final.text = nuevo

    return ET.tostring(root, encoding="utf-8")


# ======================================================================
# GENERAR DOCUMENTO COMPLETO DESDE JSON
# ======================================================================
def generar_documento(path_plantilla, estructura_json, datos_doc, path_salida):
    with zipfile.ZipFile(path_plantilla, "r") as zin:
        buffer = BytesIO()

        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():

                contenido = zin.read(item.filename)

                # Rellenar controles del documento
                if item.filename == "word/document.xml":
                    contenido = fill_content_controls(contenido, datos_doc)

                # Rellenar excels embebidos
                else:
                    for excel_info in estructura_json["excels"]:
                        if item.filename == excel_info["excel"]:
                            contenido = rellenar_excel_embebido(contenido, excel_info)

                zout.writestr(item, contenido)

        with open(path_salida, "wb") as f:
            f.write(buffer.getvalue())


# ======================================================================
# MAIN
# ======================================================================
if __name__ == "__main__":
    PLANTILLA = "ACTA_DE_REUNIÓN.docx"
    JSON_ESTRUCTURA = "estructura_acta.json"
    N = 5  # cantidad documentos a generar

    # -----------------------------------------------------------
    # NUEVO: Carpeta para los documentos generados
    # -----------------------------------------------------------
    OUTPUT_DIR = "ACTAS_GENERADAS"

    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"[OK] Carpeta creada → {OUTPUT_DIR}")
    else:
        print(f"[OK] Carpeta existente → {OUTPUT_DIR}")

    estructura_json = json.load(open(JSON_ESTRUCTURA, "r", encoding="utf-8"))

    for i in range(1, N + 1):
        datos = generar_datos_sinteticos(i, N)

        # ruta final dentro de la carpeta
        out = os.path.join(OUTPUT_DIR, f"ACTA_REUNION_{i}.docx")

        generar_documento(PLANTILLA, estructura_json, datos, out)
        print(f"[OK] Generado → {out}")

    print("\nLISTO.")

