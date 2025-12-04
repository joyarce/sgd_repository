# ===============================================================
# leer_universal_excel.py – UNIVERSAL + EJECUTABLE DESDE CONSOLA
#
# Uso:
#   python leer_universal_excel.py archivo.xlsx
#   python leer_universal_excel.py archivo.xlsx salida.json
#
# Genera JSON universal para CURVA, LISTADO HERRAMIENTAS
# y cualquier Excel nuevo.
#
# Compatible con:
#   - llenar_curva_poblamiento.py
#   - llenar_listado_herramientas.py
#
# Ahora incluye:
#   - named_ranges
#   - merged_cells
#   - cell_styles: estilos de la FILA DE ENCABEZADO de cada tabla
# ===============================================================

import json
import sys
import os
import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table


# ---------------------------------------------------------------
# UTILIDADES BÁSICAS
# ---------------------------------------------------------------

def normalize_range(r):
    return str(r)


def intersect_ranges(r1, r2):
    """Detecta si dos rangos Excel se intersectan."""
    r1 = normalize_range(r1)
    r2 = normalize_range(r2)

    min_col1, min_row1, max_col1, max_row1 = range_boundaries(r1)
    min_col2, min_row2, max_col2, max_row2 = range_boundaries(r2)

    return not (
        max_col1 < min_col2 or max_col2 < min_col1 or
        max_row1 < min_row2 or max_row2 < min_row1
    )


def infer_type(value):
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "float"
    if isinstance(value, str):
        return "formula" if value.startswith("=") else "string"
    return "string"


# ---------------------------------------------------------------
# UTILIDADES DE ESTILO
# ---------------------------------------------------------------

def _safe_color_to_str(color_obj):
    """Convierte objeto Color de openpyxl a string legible."""
    try:
        if color_obj is None:
            return None
        if getattr(color_obj, "type", None) == "rgb":
            return color_obj.rgb
        # theme / indexed / auto -> lo dejamos como str
        return str(color_obj)
    except Exception:
        return None


def get_cell_style_info(cell):
    """
    Extrae un dict con la info de estilo relevante de una celda:
    - fuente
    - alineación
    - bordes
    - relleno
    - formato numérico
    - tipo de dato
    """
    font = cell.font
    align = cell.alignment
    border = cell.border
    fill = cell.fill

    try:
        font_color = _safe_color_to_str(font.color) if font else None
    except Exception:
        font_color = None

    try:
        fill_color = _safe_color_to_str(fill.fgColor) if fill else None
    except Exception:
        fill_color = None

    style = {
        "coordinate": cell.coordinate,
        # Fuente
        "font_name": font.name if font else None,
        "font_size": float(font.size) if font and font.size else None,
        "font_bold": bool(font.bold) if font is not None else None,
        "font_italic": bool(font.italic) if font is not None else None,
        "font_color": font_color,
        # Alineación
        "alignment_horizontal": align.horizontal if align else None,
        "alignment_vertical": align.vertical if align else None,
        "alignment_wrap_text": align.wrap_text if align is not None else None,
        # Bordes
        "border_left": border.left.style if border else None,
        "border_right": border.right.style if border else None,
        "border_top": border.top.style if border else None,
        "border_bottom": border.bottom.style if border else None,
        # Relleno
        "fill_color": fill_color,
        # Formato
        "number_format": cell.number_format,
        "data_type": cell.data_type,
    }

    return style


# ---------------------------------------------------------------
# LECTOR UNIVERSAL EXCEL
# ---------------------------------------------------------------

def leer_universal_excel(path_excel, path_json_output=None):

    if not os.path.isfile(path_excel):
        raise FileNotFoundError(f"Archivo no encontrado: {path_excel}")

    wb = openpyxl.load_workbook(path_excel, data_only=False)

    data = {
        "archivo": path_excel,
        "cantidad_hojas": len(wb.worksheets),
        "named_ranges": [],
        "hojas": []
    }

    # -----------------------------------------------------------
    # Named ranges (compatible con versiones viejas y nuevas)
    # -----------------------------------------------------------
    try:
        # openpyxl viejo (<=2.6)
        defined_list = wb.defined_names.definedName
    except AttributeError:
        # openpyxl moderno (>=3.0)
        defined_list = wb.defined_names

    for dn in defined_list:
        try:
            # Caso objeto DefinedName
            name = dn.name
            ref = dn.attr_text
        except AttributeError:
            # Caso clave tipo string (dict-like)
            name = dn
            try:
                ref = wb.defined_names[name].attr_text
            except Exception:
                ref = None

        data["named_ranges"].append({
            "nombre": name,
            "rango": ref
        })

    # -----------------------------------------------------------
    # Recorrer todas las hojas
    # -----------------------------------------------------------
    for ws in wb.worksheets:

        hoja_info = {
            "nombre_hoja": ws.title,
            "tablas": [],
            "merged_cells": [str(r) for r in ws.merged_cells.ranges]
        }

        # -------------------------------------------------------
        # Tablas (ListObjects)
        # -------------------------------------------------------
        for table_name in ws.tables:

            tbl = ws.tables[table_name]

            # Seguridad: algunas plantillas dañadas pueden no devolver Table
            if not isinstance(tbl, Table):
                print(
                    f"[ADVERTENCIA] La 'tabla' '{table_name}' en hoja '{ws.title}' "
                    f"no es un objeto Table válido. Se omite."
                )
                continue

            ref = tbl.ref
            min_c, min_r, max_c, max_r = range_boundaries(ref)

            columnas = [c.name for c in tbl.tableColumns]
            filas = []

            # -------------------------
            # FILAS (solo valores)
            # -------------------------
            for row in range(min_r + 1, max_r + 1):
                fila = {}
                for i, col in enumerate(range(min_c, max_c + 1)):
                    header = columnas[i]
                    fila[header] = ws.cell(row=row, column=col).value
                filas.append(fila)

            # -------------------------
            # VALIDACIONES
            # -------------------------
            validaciones = []
            if ws.data_validations is not None:
                for dv in ws.data_validations.dataValidation:
                    for r in dv.ranges:
                        rstr = normalize_range(r)
                        if intersect_ranges(ref, rstr):

                            c1, r1, c2, r2 = range_boundaries(rstr)
                            columnas_afectadas = []

                            for c in range(c1, c2 + 1):
                                if min_c <= c <= max_c:
                                    idx = c - min_c
                                    if idx < len(columnas):
                                        columnas_afectadas.append(columnas[idx])

                            validaciones.append({
                                "tipo": dv.type,
                                "operador": dv.operator,
                                "formula1": dv.formula1,
                                "formula2": dv.formula2,
                                "permitir_nulos": dv.allowBlank,
                                "mensaje_error": dv.error,
                                "titulo_mensaje_error": dv.errorTitle,
                                "cuadro_mensaje": dv.prompt,
                                "rango_validacion": rstr,
                                "columnas_afectadas": columnas_afectadas,
                            })

            # -------------------------
            # ESTILOS DE ENCABEZADO
            # -------------------------
            cell_styles = {}
            try:
                table_range = ws[ref]
                if table_range:
                    header_row = table_range[0]   # primera fila = encabezados
                    for cell in header_row:
                        cell_styles[cell.coordinate] = get_cell_style_info(cell)
            except Exception:
                # Si por alguna razón falla, no rompemos nada
                cell_styles = {}

            # -------------------------
            # ARMAR TABLA
            # -------------------------
            hoja_info["tablas"].append({
                "nombre_tabla": table_name,
                "displayName": tbl.displayName,
                "rango": ref,
                "columnas": columnas,
                "cantidad_filas": max_r - min_r,
                "cantidad_columnas": max_c - min_c + 1,
                "filas": filas,
                "validaciones": validaciones,
                # NUEVO: estilos de encabezado por coordenada
                "cell_styles": cell_styles
            })

        data["hojas"].append(hoja_info)

    # -----------------------------------------------------------
    # Nombre de salida por defecto si no se especifica
    # -----------------------------------------------------------
    if path_json_output is None:
        base = os.path.splitext(os.path.basename(path_excel))[0]
        path_json_output = f"estructura_{base.lower()}.json"

    # -----------------------------------------------------------
    # Guardar JSON
    # -----------------------------------------------------------
    with open(path_json_output, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

    print("===============================================")
    print("  ✔ JSON GENERADO CORRECTAMENTE")
    print(f"  → {path_json_output}")
    print("===============================================")

    return data


# ---------------------------------------------------------------
# MAIN: Ejecutable desde consola
# ---------------------------------------------------------------

if __name__ == "__main__":

    if len(sys.argv) < 2:
        print("Uso:")
        print("  python leer_universal_excel.py archivo.xlsx [salida.json]")
        sys.exit(1)

    excel_path = sys.argv[1]
    salida_path = sys.argv[2] if len(sys.argv) >= 3 else None

    leer_universal_excel(excel_path, salida_path)
