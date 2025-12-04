# ===============================================================
# generar_reuniones_inicio_desde_json.py — VERSIÓN UNIVERSAL FINAL
# Consume el JSON de estructura generado por leer_universal.py y:
#   ✔ Usa controles["controles"] para definir tipos, valores, alias/tag/id
#   ✔ Rellena controles de contenido en Word (document.xml)
#   ✔ Rellena tablas Excel embebidas:
#       - RESPONSABLE → aleatorio
#       - CUMPLIMIENTO → ["SI","NO","N/A"] o lista validación
#       - OBSERVACIONES → según cumplimiento (SI/NO/N/A)
#       - % CUMP → según documento e historial (ESTADO_GLOBAL)
# Mantiene estilos y fórmulas de las tablas originales.
# ===============================================================

import os
import random
import json
from datetime import datetime, timedelta
import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO
from openpyxl import load_workbook


NS_W = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


# ===============================================================
# Catálogos profesionales
# ===============================================================

PROYECTOS = [
    "Overhaul Chancador Primario MP1000",
    "Reparación Molino SAG 40x22",
    "Cambio Coraza Molino Bolas 26x38",
    "Mantención Harnero Banana DF501",
    "Reemplazo Sistema Hidráulico Chancador Secundario HP300",
    "Upgrade de Sistema de Lubricación MP1250",
    "Servicio Major Overhaul Chancador GP500",
]

CLIENTES = [
    "BHP Escondida",
    "CODELCO Chuquicamata",
    "CODELCO Andina",
    "Collahuasi",
    "Anglo American Los Bronces",
    "Minera Centinela",
    "Minera Lomas Bayas",
]

FAENAS = [
    "Faena Escondida",
    "Faena Spence",
    "Faena Andina",
    "Faena Los Bronces",
    "Faena Radomiro Tomic",
    "Faena Collahuasi",
]

ADMINISTRADORES = [
    "Carlos Muñoz",
    "Francisca Rivas",
    "Pedro González",
    "Ana Villalobos",
    "Luis Arrieta",
    "Javiera Troncoso",
]

RESPONSABLES_OPERACION = [
    "Carlos Muñoz",
    "Francisca Rivas",
    "Pedro González",
    "Ana Villalobos",
    "Luis Arrieta",
    "Javiera Troncoso",
    "Rodrigo Pérez",
    "Marcela Olivares",
    "Cristian Valdés",
    "Patricio Carrasco",
    "Camila Rodríguez",
    "Tomás Ibáñez",
]


# ===============================================================
# Utilidades
# ===============================================================

def norm(s):
    """Normaliza textos de encabezados para evitar errores."""
    if s is None:
        return ""
    return (
        str(s)
        .strip()
        .replace("\n", "")
        .replace("\r", "")
        .upper()
    )


def get_cc_key(cc):
    """
    Aplica prioridad alias → tag → id para definir la clave con la que
    se guardará el valor en el contexto.
    """
    if cc.get("alias"):
        return cc["alias"]
    if cc.get("tag"):
        return cc["tag"]
    if cc.get("id"):
        return cc["id"]
    return None


def registrar_valor_en_contexto(ctx, cc, valor):
    """
    Registra el valor en el contexto usando TODAS las claves posibles:
    alias, tag e id, para maximizar compatibilidad.
    """
    for k in (cc.get("alias"), cc.get("tag"), cc.get("id")):
        if k:
            ctx[k] = valor


# ===============================================================
# Generación sintética para controles Word (basada en JSON)
# ===============================================================

def generar_valor_generico(alias, tipo_logico, valores=None):
    """
    Genera un valor sintético en función de:
      - alias (semántica)
      - tipo_logico (preferentemente inferred_type del JSON)
      - valores (listas de dropdown desde JSON)
    """

    alias_low = alias.lower() if alias else ""

    # Fechas
    if tipo_logico in ("date", "Date"):
        fecha = datetime.now().date() + timedelta(days=random.randint(-3, 3))
        return fecha.strftime("%Y-%m-%d")

    # Dropdown / lista
    if tipo_logico in ("dropdown", "dropDownList", "comboBox") and valores:
        valores_filtrados = [
            v for v in valores
            if v and v.strip() and v.strip().lower() not in ["elija un elemento."]
        ]
        base = valores_filtrados or valores
        return random.choice(base)

    # Semántica por alias
    if "cliente" in alias_low:
        return random.choice(CLIENTES)

    if "proyecto" in alias_low:
        return random.choice(PROYECTOS)

    if "faena" in alias_low:
        return random.choice(FAENAS)

    if "administrador" in alias_low:
        return random.choice(ADMINISTRADORES)

    if "num" in alias_low or "número" in alias_low or "numero" in alias_low:
        return f"SERV{random.randint(100000, 999999)}"

    # Genérico
    return f"Valor generado para {alias or 'campo'}"


def generar_contexto(controles, i, total_docs):
    """
    Genera un contexto de valores para TODOS los controles que vienen en el JSON.
    Usa:
      - alias / tag / id (prioridad alias → tag → id)
      - tipo (cc['tipo'])
      - inferred_type (cc['inferred_type'])
      - valores (cc['valores'])
    además de una lógica coherente de fechas de inicio/término/realización.
    """

    ctx = {}

    # Ventana temporal coherente con índice de documento
    base = datetime.now().date() - timedelta(days=total_docs + 3)
    inicio = base + timedelta(days=(i - 1))
    termino = inicio + timedelta(days=random.randint(3, 15))
    realizacion = inicio + timedelta(days=random.randint(0, (termino - inicio).days))

    for cc in controles:
        alias = cc.get("alias") or cc.get("tag") or cc.get("id") or "campo_sin_nombre"
        tipo_raw = cc.get("tipo")
        inferred = cc.get("inferred_type")
        valores = cc.get("valores")

        # Tipo lógico: primero inferred_type, luego tipo original
        tipo_logico = inferred or tipo_raw or "text"
        alias_low = alias.lower()

        # Reglas específicas de fechas por alias
        if "fecha_inicio" in alias_low:
            valor = inicio.strftime("%Y-%m-%d")
        elif "fecha_termino" in alias_low or "fecha_término" in alias_low:
            valor = termino.strftime("%Y-%m-%d")
        elif "fecha_realizacion" in alias_low or "fecha_realización" in alias_low:
            valor = realizacion.strftime("%Y-%m-%d")
        else:
            valor = generar_valor_generico(alias, tipo_logico, valores)

        # Registrar valor con las tres posibles claves: alias, tag e id
        registrar_valor_en_contexto(ctx, cc, valor)

    return ctx


# ===============================================================
# Rellenar controles Word
# ===============================================================

def fill_content_controls(document_xml_bytes, context):
    """
    Rellena los w:sdt en document.xml utilizando el contexto.
    El contexto puede tener claves por alias, tag e id; pero aquí
    usamos el tag (w:tag/@w:val) para hacer match.
    """

    root = ET.fromstring(document_xml_bytes)

    for sdt in root.findall(".//w:sdt", NS_W):
        props = sdt.find("w:sdtPr", NS_W)
        if props is None:
            continue

        tag_el = props.find("w:tag", NS_W)
        alias_el = props.find("w:alias", NS_W)
        id_el = props.find("w:id", NS_W)

        key_candidates = []

        if alias_el is not None:
            a = alias_el.get(f"{{{NS_W['w']}}}val")
            if a:
                key_candidates.append(a)

        if tag_el is not None:
            t = tag_el.get(f"{{{NS_W['w']}}}val")
            if t:
                key_candidates.append(t)

        if id_el is not None:
            cid = id_el.get(f"{{{NS_W['w']}}}val")
            if cid:
                key_candidates.append(cid)

        # Buscar la primera clave que exista en el contexto
        key = None
        for k in key_candidates:
            if k in context:
                key = k
                break

        if not key:
            continue

        contenido = sdt.find("w:sdtContent", NS_W)
        if contenido is None:
            continue

        # limpiar textos actuales
        for t in contenido.findall(".//w:t", NS_W):
            t.text = ""

        # asegurar un run
        run = contenido.find(".//w:r", NS_W)
        if run is None:
            run = ET.SubElement(contenido, f"{{{NS_W['w']}}}r")

        t_final = run.find("w:t", NS_W)
        if t_final is None:
            t_final = ET.SubElement(run, f"{{{NS_W['w']}}}t")

        t_final.text = str(context[key])

    return ET.tostring(root, encoding="utf-8")


# ===============================================================
# Rellenar Excel embebido (misma lógica que tu script actual)
# ===============================================================

OBS_SI = [
    "Actividad completada correctamente.",
    "Tarea realizada según planificación.",
    "Procedimiento ejecutado sin inconvenientes.",
    "Trabajo finalizado con conformidad del equipo.",
]

OBS_NO = [
    "Pendiente de gestión por parte del equipo.",
    "Requiere reprogramación.",
    "Faltan insumos.",
    "Detenido por condiciones en terreno.",
]

OBS_NA = [
    "No aplica para este alcance.",
    "No corresponde a esta etapa.",
]

# Estructura global para mantener consistencia de % CUMP entre documentos
ESTADO_GLOBAL = {}


def obtener_valores_validacion(ws, col_index):
    """
    Busca validaciones tipo lista en la hoja y devuelve los valores permitidos.
    Si no encuentra nada, retornará None.
    Nota: no mapeamos por columna específica, pero en tu plantilla todas
    las listas son globales y sirve igual.
    """

    permitidos = set()

    if not ws.data_validations:
        return None

    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue

        if dv.formula1 and dv.formula1.startswith('"'):
            valores = dv.formula1.strip('"').split(",")
            permitidos.update([v.strip() for v in valores])

    return list(permitidos) if permitidos else None


def rellenar_excel_embebido(xlsx_bytes, indice_doc, total_docs):
    """
    Rellena TODAS las tablas de TODOS los excels embebidos de acuerdo con:
      - RESPONSABLE: aleatorio de RESPONSABLES_OPERACION
      - CUMPLIMIENTO: ["SI", "NO", "N/A"] o lista validación
      - OBSERVACIONES: OBS_SI / OBS_NO / OBS_NA según cumplimiento
      - % CUMP: con seguimiento entre documentos usando ESTADO_GLOBAL

    Se mantiene la misma lógica exacta de tu script anterior,
    solo se mejora el nombre de la tabla usando tbl.name (más legible).
    """

    wb = load_workbook(BytesIO(xlsx_bytes), data_only=False)

    for ws in wb.worksheets:
        tablas = list(ws._tables.values())

        for t_index, tbl in enumerate(tablas, start=1):

            # Usamos el nombre real de la tabla, y si no tiene, una default.
            nombre_tabla = tbl.name or f"{ws.title}_T{t_index}"

            if nombre_tabla not in ESTADO_GLOBAL:
                ESTADO_GLOBAL[nombre_tabla] = {}

            # Rango de la tabla
            ref = tbl.ref
            celdas = ws[ref]

            # Encabezados normalizados
            raw_headers = [c.value for c in celdas[0]]
            headers = [norm(h) for h in raw_headers]

            try:
                col_item = headers.index("ITEM")        # si no existe, ValueError
                col_resp = headers.index("RESPONSABLE")
                col_cump = headers.index("CUMPLIMIENTO")
                col_obs = headers.index("OBSERVACIONES")
            except ValueError:
                # La tabla no sigue el formato esperado, se salta
                continue

            col_pct = headers.index("% CUMP.") if "% CUMP." in headers else None

            valores_validos = obtener_valores_validacion(ws, col_cump) or ["SI", "NO", "N/A"]

            # RELLENAR FILAS (excluyendo encabezado)
            for fila_idx, row in enumerate(celdas[1:], start=1):
                item_id = fila_idx  # índice de fila dentro de la tabla

                # PRIMER DOCUMENTO ---------------------------------------------
                if indice_doc == 1:
                    cumplimiento = random.choice(valores_validos)
                    responsable = random.choice(RESPONSABLES_OPERACION)

                    ESTADO_GLOBAL[nombre_tabla][item_id] = {
                        "fijo_NA": (cumplimiento == "N/A"),
                        "pct_prev": None,
                    }

                    row[col_resp].value = responsable
                    row[col_cump].value = cumplimiento

                    if cumplimiento == "NO":
                        row[col_obs].value = random.choice(OBS_NO)
                    elif cumplimiento == "N/A":
                        row[col_obs].value = random.choice(OBS_NA)
                    else:
                        row[col_obs].value = random.choice(OBS_SI)

                    if col_pct is not None:
                        if cumplimiento == "SI":
                            pct = round(1 / total_docs, 2)
                        elif cumplimiento == "NO":
                            pct = 0
                        else:  # N/A
                            pct = None

                        row[col_pct].value = pct
                        ESTADO_GLOBAL[nombre_tabla][item_id]["pct_prev"] = pct

                    continue

                # SIGUIENTES DOCUMENTOS ----------------------------------------
                memoria = ESTADO_GLOBAL[nombre_tabla].get(item_id)
                if memoria is None:
                    # Por si acaso no existía registro, se inicializa neutro
                    memoria = {"fijo_NA": False, "pct_prev": None}
                    ESTADO_GLOBAL[nombre_tabla][item_id] = memoria

                # Si el ítem quedó como N/A fijo, se mantiene así siempre
                if memoria["fijo_NA"]:
                    row[col_cump].value = "N/A"
                    row[col_resp].value = random.choice(RESPONSABLES_OPERACION)
                    row[col_obs].value = random.choice(OBS_NA)
                    if col_pct is not None:
                        row[col_pct].value = None
                        memoria["pct_prev"] = None
                    continue

                # Elegimos SI o NO (no N/A en etapas posteriores)
                valores_sin_na = [v for v in valores_validos if v.upper() != "N/A"]
                if not valores_sin_na:
                    valores_sin_na = ["SI", "NO"]
                cumplimiento = random.choice(valores_sin_na)

                responsable = random.choice(RESPONSABLES_OPERACION)
                row[col_resp].value = responsable
                row[col_cump].value = cumplimiento
                row[col_obs].value = random.choice(OBS_NO if cumplimiento == "NO" else OBS_SI)

                if col_pct is not None:
                    if cumplimiento == "NO":
                        # No avanza el % respecto al documento anterior
                        row[col_pct].value = memoria["pct_prev"]
                    else:
                        # Progreso alrededor de indice_doc / total_docs
                        progreso = round(
                            indice_doc / total_docs + random.uniform(-0.05, 0.05),
                            2,
                        )
                        progreso = max(0, min(1, progreso))
                        row[col_pct].value = progreso
                        memoria["pct_prev"] = progreso

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===============================================================
# Generar documento final
# ===============================================================

def generar_documento(path_plantilla, context, path_salida):
    """
    Copia el DOCX de plantilla y:
      - Reemplaza word/document.xml con los controles rellenados
      - Rellena todos los excels embebidos
    """

    with zipfile.ZipFile(path_plantilla, "r") as zin:
        buffer = BytesIO()

        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                contenido = zin.read(item.filename)

                if item.filename == "word/document.xml":
                    contenido = fill_content_controls(contenido, context)

                elif item.filename.startswith("word/embeddings/") and item.filename.endswith(".xlsx"):
                    contenido = rellenar_excel_embebido(
                        contenido,
                        context["__indice__"],
                        context["__total_docs__"],
                    )

                zout.writestr(item, contenido)

        with open(path_salida, "wb") as f:
            f.write(buffer.getvalue())


# ===============================================================
# MAIN
# ===============================================================

if __name__ == "__main__":
    # Archivos base
    JSON_ESTRUCTURA = "estructura_reunion_inicio.json"  # generado por leer_universal.py

    # Leemos el JSON de estructura
    with open(JSON_ESTRUCTURA, "r", encoding="utf-8") as f:
        estructura = json.load(f)

    # La plantilla se obtiene del propio JSON (campo "plantilla")
    PLANTILLA = estructura.get("plantilla", "REUNIÓN INICIO.docx")
    controles = estructura.get("controles", [])

    # N° de documentos a generar
    N_DOCUMENTOS = 5

    # Carpeta de salida
    carpeta_salida = "GENERADOS_REUNION_INICIO_JSON"
    if not os.path.exists(carpeta_salida):
        os.makedirs(carpeta_salida)
        print(f"[OK] Carpeta creada: {carpeta_salida}")
    else:
        print(f"[OK] Carpeta existente: {carpeta_salida}")

    # Reiniciamos estado global de % CUMP
    ESTADO_GLOBAL.clear()

    for i in range(1, N_DOCUMENTOS + 1):
        ctx = generar_contexto(controles, i, N_DOCUMENTOS)
        ctx["__indice__"] = i
        ctx["__total_docs__"] = N_DOCUMENTOS

        salida = os.path.join(carpeta_salida, f"REUNION_INICIO_{i}.docx")
        generar_documento(PLANTILLA, ctx, salida)

        print(f"[OK] Documento generado → {salida}")

    print("\nPROCESO COMPLETADO.")
