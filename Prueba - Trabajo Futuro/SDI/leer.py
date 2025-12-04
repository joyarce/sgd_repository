# C:\Users\jonat\Desktop\memoria\Diseño\Formatos Documentos - Final\1 LISTOS\FINAL\Pre\test\OK\SDI\leer.py (versión producción - extracción completa + corrección checkbox Word)
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from io import BytesIO
import re
import json

# Namespaces Word
NS_ALL = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "w14": "http://schemas.microsoft.com/office/word/2010/wordml",
    "w15": "http://schemas.microsoft.com/office/word/2012/wordml",
}

# Namespace Excel
NS_XL = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


# ==========================================================
# PARSEO DE styles.xml (Excel) PARA RESOLVER ESTILOS COMPLETOS
# ==========================================================
class StylesResolver:
    def __init__(self, styles_xml_text):
        self.fonts = []
        self.fills = []
        self.borders = []
        self.num_fmts = {}
        self.cell_xfs = []

        if not styles_xml_text:
            return

        root = ET.fromstring(styles_xml_text)

        numfmts_elem = root.find("s:numFmts", NS_XL)
        if numfmts_elem is not None:
            for nf in numfmts_elem.findall("s:numFmt", NS_XL):
                num_fmt_id = nf.get("numFmtId")
                format_code = nf.get("formatCode")
                if num_fmt_id:
                    self.num_fmts[int(num_fmt_id)] = format_code

        fonts_elem = root.find("s:fonts", NS_XL)
        if fonts_elem is not None:
            for f in fonts_elem.findall("s:font", NS_XL):
                name_el = f.find("s:name", NS_XL)
                sz_el = f.find("s:sz", NS_XL)
                color_el = f.find("s:color", NS_XL)
                font = {
                    "name": name_el.get("val") if name_el is not None else None,
                    "size": float(sz_el.get("val")) if sz_el is not None and sz_el.get("val") else None,
                    "bold": f.find("s:b", NS_XL) is not None,
                    "italic": f.find("s:i", NS_XL) is not None,
                    "underline": f.find("s:u", NS_XL) is not None,
                    "color": dict(color_el.attrib) if color_el is not None else None,
                }
                self.fonts.append(font)

        fills_elem = root.find("s:fills", NS_XL)
        if fills_elem is not None:
            for fl in fills_elem.findall("s:fill", NS_XL):
                pattern_el = fl.find("s:patternFill", NS_XL)
                fill = {"patternType": None, "fgColor": None, "bgColor": None}
                if pattern_el is not None:
                    fill["patternType"] = pattern_el.get("patternType")
                    fg_el = pattern_el.find("s:fgColor", NS_XL)
                    bg_el = pattern_el.find("s:bgColor", NS_XL)
                    fill["fgColor"] = dict(fg_el.attrib) if fg_el is not None else None
                    fill["bgColor"] = dict(bg_el.attrib) if bg_el is not None else None
                self.fills.append(fill)

        borders_elem = root.find("s:borders", NS_XL)
        if borders_elem is not None:
            for bd in borders_elem.findall("s:border", NS_XL):
                border = {}
                for side in ["left", "right", "top", "bottom"]:
                    side_el = bd.find(f"s:{side}", NS_XL)
                    border[side] = side_el.get("style") if side_el is not None else None
                self.borders.append(border)

        cellxfs_elem = root.find("s:cellXfs", NS_XL)
        if cellxfs_elem is not None:
            for xf in cellxfs_elem.findall("s:xf", NS_XL):
                numFmtId = xf.get("numFmtId")
                fontId = xf.get("fontId")
                fillId = xf.get("fillId")
                borderId = xf.get("borderId")

                align_el = xf.find("s:alignment", NS_XL)
                alignment = dict(align_el.attrib) if align_el is not None else None

                xf_dict = {
                    "numFmtId": int(numFmtId) if numFmtId else None,
                    "fontId": int(fontId) if fontId else None,
                    "fillId": int(fillId) if fillId else None,
                    "borderId": int(borderId) if borderId else None,
                    "alignment": alignment,
                }
                self.cell_xfs.append(xf_dict)

    def get_style_by_index(self, idx):
        if idx is None:
            return {}

        try:
            idx = int(idx)
        except:
            return {}

        if not (0 <= idx < len(self.cell_xfs)):
            return {}

        xf = self.cell_xfs[idx]
        style = {"xf_index": idx}

        numFmtId = xf.get("numFmtId")
        if numFmtId is not None:
            style["numFmtId"] = numFmtId
            style["numFmtCode"] = self.num_fmts.get(numFmtId)

        fontId = xf.get("fontId")
        if fontId is not None and 0 <= fontId < len(self.fonts):
            style["font"] = self.fonts[fontId]

        fillId = xf.get("fillId")
        if fillId is not None and 0 <= fillId < len(self.fills):
            style["fill"] = self.fills[fillId]

        borderId = xf.get("borderId")
        if borderId is not None and 0 <= borderId < len(self.borders):
            style["border"] = self.borders[borderId]

        if xf.get("alignment"):
            style["alignment"] = xf["alignment"]

        return style


# ==========================================================
# UTILIDADES WORD
# ==========================================================
def read_xml_from_docx(zip_obj, path):
    if path in zip_obj.namelist():
        return zip_obj.read(path).decode("utf-8", "ignore")
    return None


def extract_list_entries(xml_text):
    if xml_text is None:
        return []

    patterns = [
        r'<w:listEntry[^>]*w:value="(.*?)"',
        r'<w14:listEntry[^>]*w14:val="(.*?)"',
        r'<w15:listEntry[^>]*w15:val="(.*?)"',
        r'<w:listItem[^>]*w:value="(.*?)"',
        r'<w:listItem[^>]*w:displayText=".*?" w:value="(.*?)"',
    ]
    vals = []
    for p in patterns:
        vals.extend(re.findall(p, xml_text))
    return list(set(vals))


# ==========================================================
# CONTROLES DE CONTENIDO (Word + Checkbox moderno)
# ==========================================================
def extract_content_controls(docx_path):
    with zipfile.ZipFile(docx_path) as z:

        xml_files = {
            "document": read_xml_from_docx(z, "word/document.xml"),
            "styles": read_xml_from_docx(z, "word/styles.xml"),
            "settings": read_xml_from_docx(z, "word/settings.xml"),
            "glossary": read_xml_from_docx(z, "word/glossary/document.xml"),
            "numbering": read_xml_from_docx(z, "word/numbering.xml"),
        }

        valores_globales = []
        for name, xml in xml_files.items():
            vals = extract_list_entries(xml)
            valores_globales.extend(vals)
        valores_globales = list(set(valores_globales))

        root = ET.fromstring(xml_files["document"])
        controles = []

        for sdt in root.findall(".//w:sdt", NS_ALL):

            props = sdt.find("w:sdtPr", NS_ALL)
            if props is None:
                continue

            alias_node = props.find("w:alias", NS_ALL)
            tag_node = props.find("w:tag", NS_ALL)
            id_node = props.find("w:id", NS_ALL)

            alias = alias_node.get(f"{{{NS_ALL['w']}}}val") if alias_node is not None else None
            tag = tag_node.get(f"{{{NS_ALL['w']}}}val") if tag_node is not None else None
            cid = id_node.get(f"{{{NS_ALL['w']}}}val") if id_node is not None else None

            tipo = None
            valores_cc = []
            checked = None

            # DETECCIÓN DE TIPOS
            checkbox_moderno = props.find("w14:checkbox", NS_ALL)
            checkbox_antiguo = props.find("w:checkbox", NS_ALL)
            date_node = props.find("w:date", NS_ALL)
            drop = props.find("w:dropDownList", NS_ALL)
            combo = props.find("w:comboBox", NS_ALL)

            # LISTAS
            if drop is not None or combo is not None:
                tipo = "dropDownList" if drop is not None else "comboBox"

                for entry in props.findall(".//w:listEntry", NS_ALL):
                    val = entry.get(f"{{{NS_ALL['w']}}}value")
                    if val:
                        valores_cc.append(val)

                for entry in props.findall(".//w:listItem", NS_ALL):
                    val = entry.get(f"{{{NS_ALL['w']}}}value")
                    if val:
                        valores_cc.append(val)

                if not valores_cc:
                    valores_cc = valores_globales

                valores_cc = list(set(valores_cc))

            # CHECKBOX moderno
            elif checkbox_moderno is not None:
                tipo = "checkbox"

                checked_node = checkbox_moderno.find("w14:checked", NS_ALL)
                if checked_node is not None:
                    val = checked_node.get(f"{{{NS_ALL['w14']}}}val")
                    checked = (val == "1")

                valores_cc = [True, False]

            # CHECKBOX antiguo
            elif checkbox_antiguo is not None:
                tipo = "checkbox"
                val = checkbox_antiguo.get(f"{{{NS_ALL['w']}}}checked")
                if val is not None:
                    checked = (val == "1")
                valores_cc = [True, False]

            # FECHA
            elif date_node is not None:
                tipo = "date"

            # TEXTO
            else:
                tipo = "text"

            # Inferencia
            if alias and "fecha" in alias.lower():
                inferred_type = "date"
            elif tipo == "dropDownList":
                inferred_type = "dropdown"
            elif tipo == "checkbox":
                inferred_type = "bool"
            else:
                inferred_type = "text"

            controles.append({
                "alias": alias,
                "tag": tag,
                "id": cid,
                "tipo": tipo,
                "valores": valores_cc or None,
                "checked": checked if tipo == "checkbox" else None,
                "inferred_type": inferred_type,
            })

        return controles


# ==========================================================
# LECTURA DE VALOR / FÓRMULA CELDA
# ==========================================================
def get_cell_formula_or_value(cell, sheet_xml=None):
    if isinstance(cell.value, ArrayFormula):
        f = cell.value.text.strip()
        return "=" + f if not f.startswith("=") else f

    if isinstance(cell.value, str) and cell.value.startswith("="):
        return cell.value

    if sheet_xml:
        patt = rf'<c[^>]*r="{cell.coordinate}"[^>]*>.*?<f[^>]*>(.*?)</f>'
        m = re.search(patt, sheet_xml, flags=re.DOTALL)
        if m:
            f = m.group(1).strip()
            return "=" + f if not f.startswith("=") else f

    return cell.value


# ==========================================================
# FORMATO COMPLETO DE CELDA
# ==========================================================
def safe_color_to_str(color_obj):
    try:
        if color_obj is None:
            return None
        if getattr(color_obj, "type", None) == "rgb":
            return color_obj.rgb
        return str(color_obj)
    except:
        return None


def get_complete_cell_format(cell, sheet_xml=None, styles_resolver=None):
    try:
        font_color = safe_color_to_str(cell.font.color) if cell.font else None
    except:
        font_color = None

    try:
        fill_color = safe_color_to_str(cell.fill.fgColor) if cell.fill else None
    except:
        fill_color = None

    fmt = {
        "coordinate": cell.coordinate,
        "font_name": cell.font.name if cell.font else None,
        "font_size": float(cell.font.size) if cell.font and cell.font.size else None,
        "font_bold": bool(cell.font.bold) if cell.font else None,
        "font_italic": bool(cell.font.italic) if cell.font else None,
        "font_color": font_color,
        "alignment_horizontal": cell.alignment.horizontal if cell.alignment else None,
        "alignment_vertical": cell.alignment.vertical if cell.alignment else None,
        "border_left": cell.border.left.style if cell.border else None,
        "border_right": cell.border.right.style if cell.border else None,
        "border_top": cell.border.top.style if cell.border else None,
        "border_bottom": cell.border.bottom.style if cell.border else None,
        "fill_color": fill_color,
        "number_format": cell.number_format,
        "data_type": cell.data_type,
        "is_formula": isinstance(cell.value, str) and cell.value.startswith("="),
    }

    xml_info = None
    if sheet_xml:
        xml_info = get_cell_xml_type(sheet_xml, cell.coordinate)
        fmt.update({
            "xml_t": xml_info.get("t"),
            "xml_s": xml_info.get("s"),
            "xml_is_formula": xml_info.get("is_formula"),
            "xml_raw_attrs": xml_info.get("style_attrs", {}),
        })

    if styles_resolver and xml_info:
        s_idx = xml_info.get("s")
        if s_idx is not None:
            resolved = styles_resolver.get_style_by_index(s_idx)
            fmt["resolved_style"] = resolved

    return fmt


def get_cell_xml_type(sheet_xml, cell_ref):
    patt = rf'<c[^>]*r="{cell_ref}"([^>]*)>(.*?)</c>'
    m = re.search(patt, sheet_xml, flags=re.DOTALL)
    if not m:
        return {"t": None, "s": None, "is_formula": False, "style_attrs": {}}

    attrs = m.group(1)
    inner = m.group(2)
    tipo = None
    estilo = None
    is_formula = "<f" in inner

    mt = re.search(r't="(.*?)"', attrs)
    if mt:
        tipo = mt.group(1)

    ms = re.search(r's="(.*?)"', attrs)
    if ms:
        estilo = ms.group(1)

    style_attrs = {k: v for k, v in re.findall(r'(\w+)="([^"]*)"', attrs)}

    return {
        "t": tipo,
        "s": estilo,
        "is_formula": is_formula,
        "style_attrs": style_attrs,
    }


# ==========================================================
# EXTRACCIÓN DE EXCEL EMBEBIDOS
# ==========================================================
def extract_embedded_excel(docx_path):
    with zipfile.ZipFile(docx_path) as z:

        excel_files = [
            f for f in z.namelist()
            if f.startswith("word/embeddings/") and f.endswith(".xlsx")
        ]

        results = []

        for ef in excel_files:

            content = z.read(ef)
            wb = load_workbook(BytesIO(content), data_only=False)

            sheet_xmls = {}
            styles_xml_text = None

            with zipfile.ZipFile(BytesIO(content)) as xz:
                for item in xz.namelist():
                    if item.startswith("xl/worksheets/sheet") and item.endswith(".xml"):
                        key = item.split("/")[-1].replace(".xml", "")
                        sheet_xmls[key] = xz.read(item).decode("utf-8", "ignore")
                    if item == "xl/styles.xml":
                        styles_xml_text = xz.read(item).decode("utf-8", "ignore")

            styles_resolver = StylesResolver(styles_xml_text)

            excel_data = {
                "excel": ef,
                "tablas": [],
                "nombres_definidos": []
            }

            for name, defn in wb.defined_names.items():
                for ws_name, cell_ref in list(defn.destinations):
                    ws = wb[ws_name.replace("'", "")]
                    xml = sheet_xmls.get(f"sheet{ws._id}", "")

                    try:
                        celdas = ws[cell_ref]
                        if hasattr(celdas, "value"):
                            val = get_cell_formula_or_value(celdas, xml)
                        else:
                            val = [
                                [get_cell_formula_or_value(c, xml) for c in fila]
                                for fila in celdas
                            ]
                    except:
                        val = None

                    excel_data["nombres_definidos"].append({
                        "nombre": name,
                        "ref": f"{ws_name}!{cell_ref}",
                        "valor": val
                    })

            for ws in wb.worksheets:
                xml = sheet_xmls.get(f"sheet{ws._id}", "")
                for tbl in ws._tables.values():

                    ref = tbl.ref
                    rows = []
                    cell_styles = {}

                    excel_range = ws[ref]

                    if excel_range:
                        first_row = excel_range[0]
                        for c in first_row:
                            cell_styles[c.coordinate] = get_complete_cell_format(
                                c, sheet_xml=xml, styles_resolver=styles_resolver
                            )

                    for row in excel_range:
                        rows.append([get_cell_formula_or_value(c, xml) for c in row])

                    excel_data["tablas"].append({
                        "worksheet": ws.title,
                        "tabla": tbl.name,
                        "rango": ref,
                        "columnas": rows[0] if rows else [],
                        "registros": rows[1:] if len(rows) > 1 else [],
                        "cell_styles": cell_styles,
                    })

            results.append(excel_data)

        return results


# ==========================================================
# GENERAR JSON COMPLETO
# ==========================================================
def generar_json_estructura(docx_path, json_output="estructura_sdi.json"):
    controles = extract_content_controls(docx_path)
    excels = extract_embedded_excel(docx_path)

    estructura = {
        "plantilla": docx_path,
        "controles": controles,
        "excels": excels,
    }

    with open(json_output, "w", encoding="utf-8") as f:
        json.dump(estructura, f, indent=4, ensure_ascii=False)

    print(f"\n[OK] JSON generado correctamente → {json_output}\n")


# ==========================================================
# MAIN
# ==========================================================
if __name__ == "__main__":

    docx_path = "SDI.docx"

    print("\n=== Generando estructura JSON completa ===")
    generar_json_estructura(docx_path)

    print("\n=== CONTROLES DE CONTENIDO DETECTADOS ===")
    for cc in extract_content_controls(docx_path):
        print(cc)

    print("\n=== EXCEL EMBEBIDOS DETECTADOS ===")
    for ex in extract_embedded_excel(docx_path):
        print("\nArchivo:", ex["excel"])

        print("\n--- Nombres Definidos ---")
        for nd in ex["nombres_definidos"]:
            print(nd)

        print("\n--- Tablas ---")
        for t in ex["tablas"]:
            print(t)
