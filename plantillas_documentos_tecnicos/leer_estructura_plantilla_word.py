# ======================================================================
# leer_estructura_plantilla_word.py — VERSIÓN COMPATIBLE DEFINITIVA (2025)
# - Alineado al backend nuevo (tipo_detalle, comparar_estructuras, versionado)
# - Retorna SIEMPRE:
#       controles:          lista total unificada (doc + header + footer)
#       tablas_word:        tablas Word nativas normalizadas
#       excels:             excels embebidos
#       imagenes:           archivos media
#       metadata:           TODO lo demás (parrafos, campos, origenes, etc.)
# ======================================================================

import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from io import BytesIO
import re
import json
import os
import datetime

# Namespaces Word
NS_ALL = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "w14": "http://schemas.microsoft.com/office/word/2010/wordml",
    "w15": "http://schemas.microsoft.com/office/word/2012/wordml",
}

NS_W = {"w": NS_ALL["w"]}

# Namespace Excel
NS_XL = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


# ======================================================================
# FIRMA ESTRUCTURAL — SE MANTIENE IGUAL
# ======================================================================
def extract_structural_signature(estructura: dict) -> dict:
    if not isinstance(estructura, dict):
        return {}

    firma = {
        "controles": [],
        "tablas_word": [],
        "excels": []
    }

    # --- CONTROLES
    controles = estructura.get("controles", [])
    if isinstance(controles, list):
        for c in controles:
            if isinstance(c, dict):
                firma["controles"].append({
                    "tipo": c.get("tipo"),
                    "alias": c.get("alias"),
                    "tag": c.get("tag")
                })

    # --- TABLAS WORD (NORMALIZADAS)
    tablas_word = estructura.get("tablas_word", [])
    if isinstance(tablas_word, list):
        for t in tablas_word:
            if isinstance(t, dict):
                firma["tablas_word"].append({
                    "n_filas": t.get("n_filas", 0),
                    "n_columnas": t.get("n_columnas", 0)
                })

    # --- EXCELS
    excels = estructura.get("excels", [])
    if isinstance(excels, list):
        for e in excels:
            if isinstance(e, dict):
                firma["excels"].append({
                    "excel": e.get("excel"),
                    "n_tablas": len(e.get("tablas", []))
                })

    # Orden determinista
    firma["controles"] = sorted(
        firma["controles"],
        key=lambda x: (x.get("tipo") or "", x.get("alias") or "")
    )
    firma["tablas_word"] = sorted(
        firma["tablas_word"],
        key=lambda x: (x.get("n_filas", 0), x.get("n_columnas", 0))
    )
    firma["excels"] = sorted(
        firma["excels"],
        key=lambda x: x.get("excel") or ""
    )

    return firma


# ======================================================================
# JSON-SAFE
# ======================================================================
def json_safe(obj):
    if isinstance(obj, (datetime.datetime, datetime.date)):
        return obj.isoformat()
    if isinstance(obj, bytes):
        return obj.decode("utf-8", errors="ignore")
    return str(obj)


# ======================================================================
# WORD HELPERS
# ======================================================================
def read_xml_from_docx(zip_obj, path):
    if path in zip_obj.namelist():
        return zip_obj.read(path).decode("utf-8", "ignore")
    return None


def extract_list_entries(xml_text):
    if xml_text is None:
        return []
    patterns = [
        r'<w:listEntry[^>]*w:value="(.*?)"',
        r'<w14:listEntry[^>]*w14:val="(.*?)"',
        r'<w15:listEntry[^>]*w15:val="(.*?)"',
        r'<w:listItem[^>]*w:value="(.*?)"',
        r'<w:listItem[^>]*w:displayText=".*?" w:value="(.*?)"',
    ]
    vals = []
    for p in patterns:
        vals.extend(re.findall(p, xml_text))
    return list(set(vals))


# ======================================================================
# CONTROLES DE CONTENIDO
# ======================================================================
def _extract_sdt_list_from_root(root, valores_globales):
    controles = []

    for sdt in root.findall(".//w:sdt", NS_ALL):

        props = sdt.find("w:sdtPr", NS_ALL)
        if props is None:
            continue

        alias_node = props.find("w:alias", NS_ALL)
        tag_node = props.find("w:tag", NS_ALL)
        id_node = props.find("w:id", NS_ALL)

        alias = alias_node.get(f"{{{NS_ALL['w']}}}val") if alias_node is not None else None
        tag = tag_node.get(f"{{{NS_ALL['w']}}}val") if tag_node is not None else None
        cid = id_node.get(f"{{{NS_ALL['w']}}}val") if id_node is not None else None

        tipo = None
        valores_cc = []
        checked = None

        checkbox_moderno = props.find("w14:checkbox", NS_ALL)
        checkbox_antiguo = props.find("w:checkbox", NS_ALL)
        date_node = props.find("w:date", NS_ALL)
        drop = props.find("w:dropDownList", NS_ALL)
        combo = props.find("w:comboBox", NS_ALL)
        picture = props.find("w:picture", NS_ALL)
        richtext = props.find("w:richText", NS_ALL)
        repeating_section = props.find("w:repeatingSection", NS_ALL)
        repeating_item = props.find("w:repeatingSectionItem", NS_ALL)

        # Dropdown
        if drop is not None or combo is not None:
            tipo = "dropDownList" if drop is not None else "comboBox"
            for entry in props.findall(".//w:listEntry", NS_ALL):
                val = entry.get(f"{{{NS_ALL['w']}}}value")
                if val:
                    valores_cc.append(val)
            for entry in props.findall(".//w:listItem", NS_ALL):
                val = entry.get(f"{{{NS_ALL['w']}}}value")
                if val:
                    valores_cc.append(val)

            if not valores_cc:
                valores_cc = valores_globales

            valores_cc = list(set(valores_cc))

        # Checkbox
        elif checkbox_moderno is not None:
            tipo = "checkbox"
            checked_node = checkbox_moderno.find("w14:checked", NS_ALL)
            if checked_node is not None:
                checked = checked_node.get(f"{{{NS_ALL['w14']}}}val") == "1"
            valores_cc = [True, False]

        elif checkbox_antiguo is not None:
            tipo = "checkbox"
            val = checkbox_antiguo.get(f"{{{NS_ALL['w']}}}checked")
            if val is not None:
                checked = val == "1"
            valores_cc = [True, False]

        elif date_node is not None:
            tipo = "date"

        elif picture is not None:
            tipo = "picture"

        elif repeating_section is not None:
            tipo = "repeatingSection"

        elif repeating_item is not None:
            tipo = "repeatingSectionItem"

        elif richtext is not None:
            tipo = "richText"

        else:
            tipo = "text"

        # Inferencia
        if alias and isinstance(alias, str) and "fecha" in alias.lower():
            inferred_type = "date"
        elif tipo == "dropDownList":
            inferred_type = "dropdown"
        elif tipo == "checkbox":
            inferred_type = "bool"
        else:
            inferred_type = "text"

        # Valor actual
        texto_actual = []
        for t in sdt.findall(".//w:sdtContent//w:t", NS_ALL):
            if t.text:
                texto_actual.append(t.text)
        texto_actual = "".join(texto_actual).strip() if texto_actual else None

        controles.append(
            {
                "alias": alias,
                "tag": tag,
                "id": cid,
                "tipo": tipo,
                "valores": valores_cc or None,
                "checked": checked if tipo == "checkbox" else None,
                "inferred_type": inferred_type,
                "texto_actual": texto_actual,
            }
        )

    return controles


def extract_content_controls_document(docx_path):
    with zipfile.ZipFile(docx_path) as z:

        xml_files = {
            "document": read_xml_from_docx(z, "word/document.xml"),
            "styles": read_xml_from_docx(z, "word/styles.xml"),
            "settings": read_xml_from_docx(z, "word/settings.xml"),
            "glossary": read_xml_from_docx(z, "word/glossary/document.xml"),
            "numbering": read_xml_from_docx(z, "word/numbering.xml"),
        }

        valores_globales = []
        for xml in xml_files.values():
            valores_globales.extend(extract_list_entries(xml))
        valores_globales = list(set(valores_globales))

        root_doc = ET.fromstring(xml_files["document"])
        return _extract_sdt_list_from_root(root_doc, valores_globales)


def extract_content_controls_headers_footers(docx_path):

    controles_header = []
    controles_footer = []

    with zipfile.ZipFile(docx_path) as z:

        xml_global_candidates = [
            read_xml_from_docx(z, "word/styles.xml"),
            read_xml_from_docx(z, "word/settings.xml"),
            read_xml_from_docx(z, "word/numbering.xml"),
            read_xml_from_docx(z, "word/glossary/document.xml"),
        ]

        valores_globales = []
        for xml in xml_global_candidates:
            valores_globales.extend(extract_list_entries(xml))
        valores_globales = list(set(valores_globales))

        for name in z.namelist():

            # HEADERS
            if name.startswith("word/header") and name.endswith(".xml"):
                xml = read_xml_from_docx(z, name)
                if xml:
                    try:
                        root = ET.fromstring(xml)
                        sdt_list = _extract_sdt_list_from_root(root, valores_globales)
                        for c in sdt_list:
                            c["origen"] = os.path.basename(name)
                        controles_header.extend(sdt_list)
                    except Exception:
                        pass

            # FOOTERS
            if name.startswith("word/footer") and name.endswith(".xml"):
                xml = read_xml_from_docx(z, name)
                if xml:
                    try:
                        root = ET.fromstring(xml)
                        sdt_list = _extract_sdt_list_from_root(root, valores_globales)
                        for c in sdt_list:
                            c["origen"] = os.path.basename(name)
                        controles_footer.extend(sdt_list)
                    except Exception:
                        pass

    return controles_header, controles_footer


# ======================================================================
# TABLAS NATIVAS WORD — NORMALIZADAS
# ======================================================================
def extract_word_tables(docx_path):
    tablas = []

    with zipfile.ZipFile(docx_path) as z:

        xml_doc = read_xml_from_docx(z, "word/document.xml")
        if not xml_doc:
            return tablas

        root = ET.fromstring(xml_doc)

        # Enumeración de tablas Word nativas
        for i, tbl in enumerate(root.findall(".//w:tbl", NS_ALL), start=1):

            filas = []

            # Cada fila <w:tr>
            for r in tbl.findall("w:tr", NS_ALL):

                celdas = []

                # Cada celda <w:tc>
                for c in r.findall("w:tc", NS_ALL):

                    textos = []

                    for t in c.findall(".//w:t", NS_ALL):
                        if t.text:
                            textos.append(t.text)

                    celdas.append("".join(textos).strip())

                filas.append(celdas)

            # NORMALIZACIÓN
            n_filas = len(filas)
            n_columnas = len(filas[0]) if filas else 0

            tablas.append(
                {
                    "index": i,
                    "filas": filas,
                    "n_filas": n_filas,
                    "n_columnas": n_columnas,
                    "origen": "document",
                }
            )

    return tablas



# ======================================================================
# EXCEL EMBEBIDOS (estructura profunda)
# ======================================================================
def extract_embedded_excel(docx_path):
    results = []

    with zipfile.ZipFile(docx_path) as z:

        excel_files = [
            f for f in z.namelist()
            if f.startswith("word/embeddings/") and f.endswith(".xlsx")
        ]

        for ef in excel_files:

            content = z.read(ef)

            try:
                wb = load_workbook(BytesIO(content), data_only=False)
            except Exception:
                continue

            excel_data = {
                "excel": ef,
                "tablas": [],
                "nombres_definidos": []
            }

            # Tablas en el Excel embebido
            for ws in wb.worksheets:

                for tbl in ws._tables.values():

                    ref = tbl.ref
                    excel_range = ws[ref]

                    filas = []
                    for row in excel_range:
                        filas.append([c.value for c in row])

                    excel_data["tablas"].append(
                        {
                            "worksheet": ws.title,
                            "tabla": tbl.name,
                            "rango": ref,
                            "filas": filas,
                            "n_filas": len(filas),
                            "n_columnas": len(filas[0]) if filas else 0
                        }
                    )

            results.append(excel_data)

    return results


# ======================================================================
# IMÁGENES WORD
# ======================================================================
def extract_images_info(docx_path):

    imagenes = []

    with zipfile.ZipFile(docx_path) as z:

        content_types_xml = read_xml_from_docx(z, "[Content_Types].xml")
        type_map = {}

        if content_types_xml:
            root_ct = ET.fromstring(content_types_xml)
            for override in root_ct.findall(
                ".//{http://schemas.openxmlformats.org/package/2006/content-types}Override"
            ):
                name = override.get("PartName")
                ctype = override.get("ContentType")
                if "/media/" in str(name):
                    type_map[name] = ctype

        for name in z.namelist():
            if name.startswith("word/media/"):
                part_name = "/" + name
                imagenes.append(
                    {
                        "nombre": os.path.basename(name),
                        "ruta": name,
                        "content_type": type_map.get(part_name),
                    }
                )

    return imagenes


# ======================================================================
# CAMPOS Word
# ======================================================================
def extract_word_fields(docx_path):

    campos = []

    with zipfile.ZipFile(docx_path) as z:
        xml_doc = read_xml_from_docx(z, "word/document.xml")
        if not xml_doc:
            return campos

        root = ET.fromstring(xml_doc)

        for fld in root.findall(".//w:fldSimple", NS_ALL):
            instr = fld.get(f"{{{NS_ALL['w']}}}instr")
            textos = [t.text for t in fld.findall(".//w:t", NS_ALL) if t.text]
            campos.append(
                {
                    "tipo": "fldSimple",
                    "instr": instr,
                    "texto": "".join(textos).strip() if textos else None,
                }
            )

        for instr in root.findall(".//w:instrText", NS_ALL):
            txt = (instr.text or "").strip()
            if txt:
                campos.append(
                    {
                        "tipo": "instrText",
                        "instr": txt,
                    }
                )

    return campos


# ======================================================================
# PÁRRAFOS
# ======================================================================
def extract_paragraphs(docx_path, max_chars_por_parrafo=500):

    parrafos = []

    with zipfile.ZipFile(docx_path) as z:
        xml_doc = read_xml_from_docx(z, "word/document.xml")
        if not xml_doc:
            return parrafos

        root = ET.fromstring(xml_doc)

        for i, p in enumerate(root.findall(".//w:p", NS_ALL), start=1):

            ppr = p.find("w:pPr", NS_ALL)
            pstyle = None

            if ppr is not None:
                st = ppr.find("w:pStyle", NS_ALL)
                if st is not None:
                    pstyle = st.get(f"{{{NS_ALL['w']}}}val")

            textos = [t.text for t in p.findall(".//w:t", NS_ALL) if t.text]
            txt = "".join(textos).strip()

            if len(txt) > max_chars_por_parrafo:
                txt = txt[:max_chars_por_parrafo] + "..."

            if txt:
                parrafos.append(
                    {
                        "index": i,
                        "estilo": pstyle,
                        "texto": txt,
                    }
                )

    return parrafos

# ======================================================================
# SANITIZADOR JSON UNIVERSAL — ELIMINA ArrayFormula y cualquier objeto
# ======================================================================

def json_sanitize_deep(obj):
    """
    Convierte una estructura compleja (dict/list) a una estructura 100%
    JSON-serializable sin perder información.
    """
    # Tipos básicos OK
    if isinstance(obj, (str, int, float, bool)) or obj is None:
        return obj

    # Fecha/hora
    if isinstance(obj, (datetime.datetime, datetime.date)):
        return obj.isoformat()

    # Bytes → string
    if isinstance(obj, bytes):
        return obj.decode("utf-8", errors="ignore")

    # ArrayFormula u otros objetos no serializables
    if isinstance(obj, ArrayFormula):
        return str(obj)  # normalmente devuelve algo como "{=A1+A2}"

    # Diccionario
    if isinstance(obj, dict):
        return {k: json_sanitize_deep(v) for k, v in obj.items()}

    # Iterables (listas, tuplas, sets…)
    if isinstance(obj, (list, tuple, set)):
        return [json_sanitize_deep(v) for v in obj]

    # CUALQUIER OTRO TIPO → convertir a str
    return str(obj)

# ======================================================================
# 🚀 FUNCIÓN PRINCIPAL — ¡VERSIÓN DEFINITIVA 2025!
# ======================================================================
def generar_estructura(docx_path):
    """
    ESTA ES LA FUNCIÓN QUE SE GUARDA EN BD.
    Y LA QUE SE USA PARA VALIDAR DOCUMENTOS SUBIDOS.
    """

    # 1. Controles
    controles_doc = extract_content_controls_document(docx_path)
    controles_header, controles_footer = extract_content_controls_headers_footers(docx_path)

    # Unificación
    controles_unificados = [
        *controles_doc,
        *controles_header,
        *controles_footer,
    ]

    # 2. Tablas Word — normalizadas (NUNCA faltan n_filas/n_columnas)
    tablas_word = extract_word_tables(docx_path)

    # 3. Excel embebidos
    excels = extract_embedded_excel(docx_path)

    # 4. Imágenes
    imagenes = extract_images_info(docx_path)

    # 5. Campos Word
    campos = extract_word_fields(docx_path)

    # 6. Párrafos
    parrafos = extract_paragraphs(docx_path)

    # 7. Metadata
    metadata = {
        "plantilla": os.path.basename(docx_path),
        "controles_header": controles_header,
        "controles_footer": controles_footer,
        "campos_word": campos,
        "parrafos": parrafos,
    }

    # 8. RETORNO FINAL — FORMATO ESTÁNDAR
    estructura_final =  {
        "controles": controles_unificados,
        "tablas_word": tablas_word,
        "excels": excels,
        "imagenes": imagenes,
        "metadata": metadata,
    }

    return json_sanitize_deep(estructura_final)   

