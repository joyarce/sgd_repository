# leer.py (versión producción - extracción completa de estilos primera fila)
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from io import BytesIO
import re

# Namespaces Word
NS_ALL = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "w14": "http://schemas.microsoft.com/office/word/2010/wordml",
    "w15": "http://schemas.microsoft.com/office/word/2012/wordml"
}

# Namespace Excel
NS_XL = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


# ==========================================================
# PARSEO DE styles.xml (Excel) PARA RESOLVER ESTILOS COMPLETOS
# ==========================================================
class StylesResolver:
    def __init__(self, styles_xml_text):
        self.fonts = []
        self.fills = []
        self.borders = []
        self.num_fmts = {}
        self.cell_xfs = []

        if not styles_xml_text:
            return

        root = ET.fromstring(styles_xml_text)

        # numFmts (formatos numéricos personalizados)
        numfmts_elem = root.find("s:numFmts", NS_XL)
        if numfmts_elem is not None:
            for nf in numfmts_elem.findall("s:numFmt", NS_XL):
                num_fmt_id = nf.get("numFmtId")
                format_code = nf.get("formatCode")
                if num_fmt_id:
                    self.num_fmts[int(num_fmt_id)] = format_code

        # fonts
        fonts_elem = root.find("s:fonts", NS_XL)
        if fonts_elem is not None:
            for f in fonts_elem.findall("s:font", NS_XL):
                name_el = f.find("s:name", NS_XL)
                sz_el = f.find("s:sz", NS_XL)
                color_el = f.find("s:color", NS_XL)
                font = {
                    "name": name_el.get("val") if name_el is not None else None,
                    "size": float(sz_el.get("val")) if sz_el is not None and sz_el.get("val") else None,
                    "bold": f.find("s:b", NS_XL) is not None,
                    "italic": f.find("s:i", NS_XL) is not None,
                    "underline": f.find("s:u", NS_XL) is not None,
                    "color": dict(color_el.attrib) if color_el is not None else None,
                }
                self.fonts.append(font)

        # fills
        fills_elem = root.find("s:fills", NS_XL)
        if fills_elem is not None:
            for fl in fills_elem.findall("s:fill", NS_XL):
                pattern_el = fl.find("s:patternFill", NS_XL)
                fill = {"patternType": None, "fgColor": None, "bgColor": None}
                if pattern_el is not None:
                    fill["patternType"] = pattern_el.get("patternType")
                    fg_el = pattern_el.find("s:fgColor", NS_XL)
                    bg_el = pattern_el.find("s:bgColor", NS_XL)
                    fill["fgColor"] = dict(fg_el.attrib) if fg_el is not None else None
                    fill["bgColor"] = dict(bg_el.attrib) if bg_el is not None else None
                self.fills.append(fill)

        # borders
        borders_elem = root.find("s:borders", NS_XL)
        if borders_elem is not None:
            for bd in borders_elem.findall("s:border", NS_XL):
                border = {}
                for side in ["left", "right", "top", "bottom"]:
                    side_el = bd.find(f"s:{side}", NS_XL)
                    border[side] = side_el.get("style") if side_el is not None else None
                self.borders.append(border)

        # cellXfs (estilos aplicados a celdas)
        cellxfs_elem = root.find("s:cellXfs", NS_XL)
        if cellxfs_elem is not None:
            for xf in cellxfs_elem.findall("s:xf", NS_XL):
                numFmtId = xf.get("numFmtId")
                fontId = xf.get("fontId")
                fillId = xf.get("fillId")
                borderId = xf.get("borderId")

                align_el = xf.find("s:alignment", NS_XL)
                alignment = None
                if align_el is not None:
                    alignment = dict(align_el.attrib)

                xf_dict = {
                    "numFmtId": int(numFmtId) if numFmtId is not None else None,
                    "fontId": int(fontId) if fontId is not None else None,
                    "fillId": int(fillId) if fillId is not None else None,
                    "borderId": int(borderId) if borderId is not None else None,
                    "alignment": alignment,
                }
                self.cell_xfs.append(xf_dict)

    def get_style_by_index(self, s_idx):
        """
        s_idx es el índice 's' de la celda (atributo s="N" en <c>).
        Devuelve un dict con font, fill, border, numFmt y alignment.
        """
        if s_idx is None:
            return {}

        try:
            idx = int(s_idx)
        except Exception:
            return {}

        if idx < 0 or idx >= len(self.cell_xfs):
            return {}

        xf = self.cell_xfs[idx]
        style = {"xf_index": idx}

        # numFmt
        numFmtId = xf.get("numFmtId")
        if numFmtId is not None:
            style["numFmtId"] = numFmtId
            style["numFmtCode"] = self.num_fmts.get(numFmtId)

        # font
        fontId = xf.get("fontId")
        if fontId is not None and 0 <= fontId < len(self.fonts):
            style["font"] = self.fonts[fontId]

        # fill
        fillId = xf.get("fillId")
        if fillId is not None and 0 <= fillId < len(self.fills):
            style["fill"] = self.fills[fillId]

        # border
        borderId = xf.get("borderId")
        if borderId is not None and 0 <= borderId < len(self.borders):
            style["border"] = self.borders[borderId]

        # alignment
        if xf.get("alignment") is not None:
            style["alignment"] = xf["alignment"]

        return style


# ==========================================================
# INFO XML DE CELDA (tipo, estilo, etc. desde sheetX.xml)
# ==========================================================
def get_cell_xml_type(sheet_xml, cell_ref):
    """
    Retorna info básica de la celda según el XML <c> de la hoja:
    tipo (t), estilo (s), si tiene fórmula, y otros atributos crudos.
    """
    patt = rf'<c[^>]*r="{cell_ref}"([^>]*)>(.*?)</c>'
    m = re.search(patt, sheet_xml, flags=re.DOTALL)
    if not m:
        return {"t": None, "s": None, "is_formula": False, "style_attrs": {}}

    attrs = m.group(1)
    inner = m.group(2)

    tipo = None
    estilo = None
    is_formula = "<f" in inner

    # Buscar t="..."
    mt = re.search(r't="(.*?)"', attrs)
    if mt:
        tipo = mt.group(1)

    # Buscar s="..."
    ms = re.search(r's="(.*?)"', attrs)
    if ms:
        estilo = ms.group(1)

    # Extraer otros atributos (lo que venga en <c>)
    style_attrs = {k: v for k, v in re.findall(r'(\w+)="([^"]*)"', attrs)}

    return {
        "t": tipo,
        "s": estilo,
        "is_formula": is_formula,
        "style_attrs": style_attrs,
    }


# ==========================================================
# FORMATO COMPLETO DE UNA CELDA
# (openpyxl + XML de hoja + styles.xml)
# ==========================================================
def safe_color_to_str(color_obj):
    """Convierte color de openpyxl a string segura."""
    try:
        if color_obj is None:
            return None
        if getattr(color_obj, "type", None) == "rgb":
            return color_obj.rgb
        # theme/indexed no lo resolvemos aquí; podríamos agregar lógica extra
        return str(color_obj)
    except Exception:
        return None


def get_complete_cell_format(cell, sheet_xml=None, styles_resolver=None):
    """
    Devuelve un diccionario con TODO el estilo disponible de una celda:
    - openpyxl: font, fill, alignment, borders, number_format, data_type
    - XML <c>: t, s, attrs, is_formula
    - styles.xml: font/fill/border/numFmt resueltos por s
    """

    # --- Nivel openpyxl ---
    try:
        font_color = safe_color_to_str(cell.font.color) if cell.font else None
    except Exception:
        font_color = None

    try:
        fill_color = safe_color_to_str(cell.fill.fgColor) if cell.fill else None
    except Exception:
        fill_color = None

    fmt = {
        "coordinate": cell.coordinate,
        "font_name": cell.font.name if cell.font else None,
        "font_size": float(cell.font.size) if cell.font and cell.font.size else None,
        "font_bold": bool(cell.font.bold) if cell.font else None,
        "font_italic": bool(cell.font.italic) if cell.font else None,
        "font_color": font_color,
        "alignment_horizontal": cell.alignment.horizontal if cell.alignment else None,
        "alignment_vertical": cell.alignment.vertical if cell.alignment else None,
        "border_left": cell.border.left.style if cell.border else None,
        "border_right": cell.border.right.style if cell.border else None,
        "border_top": cell.border.top.style if cell.border else None,
        "border_bottom": cell.border.bottom.style if cell.border else None,
        "fill_color": fill_color,
        "number_format": cell.number_format,
        "data_type": cell.data_type,
        "is_formula": isinstance(cell.value, str) and cell.value.startswith("="),
    }

    # --- Nivel XML de hoja (sheetX.xml) ---
    xml_info = None
    if sheet_xml:
        xml_info = get_cell_xml_type(sheet_xml, cell.coordinate)
        fmt.update({
            "xml_t": xml_info.get("t"),
            "xml_s": xml_info.get("s"),
            "xml_is_formula": xml_info.get("is_formula"),
            "xml_raw_attrs": xml_info.get("style_attrs", {}),
        })

    # --- Nivel styles.xml (resolver por índice s) ---
    if styles_resolver is not None and xml_info is not None:
        s_idx = xml_info.get("s")
        if s_idx is not None:
            resolved = styles_resolver.get_style_by_index(s_idx)
            fmt["resolved_style"] = resolved

    return fmt


# ==========================================================
# UTILIDADES WORD
# ==========================================================
def read_xml_from_docx(zip_obj, path):
    if path in zip_obj.namelist():
        return zip_obj.read(path).decode("utf-8", "ignore")
    return None


def extract_list_entries(xml_text):
    if xml_text is None:
        return []
    patterns = [
        r'<w:listEntry[^>]*w:value="(.*?)"',
        r'<w14:listEntry[^>]*w14:val="(.*?)"',
        r'<w15:listEntry[^>]*w15:val="(.*?)"',
        r'<w:listItem[^>]*w:value="(.*?)"',
        r'<w:listItem[^>]*w:displayText=".*?" w:value="(.*?)"'
    ]
    vals = []
    for p in patterns:
        vals.extend(re.findall(p, xml_text))
    return list(set(vals))


# ==========================================================
# CONTROLES DE CONTENIDO EN WORD
# ==========================================================
def extract_content_controls(docx_path):
    with zipfile.ZipFile(docx_path) as z:

        xml_files = {
            "document": read_xml_from_docx(z, "word/document.xml"),
            "styles": read_xml_from_docx(z, "word/styles.xml"),
            "settings": read_xml_from_docx(z, "word/settings.xml"),
            "glossary": read_xml_from_docx(z, "word/glossary/document.xml"),
            "numbering": read_xml_from_docx(z, "word/numbering.xml")
        }

        # Valores globales
        valores_globales = []
        for name, xml in xml_files.items():
            vals = extract_list_entries(xml)
            if vals:
                print(f"[+] Valores encontrados en {name}.xml:", vals)
            valores_globales.extend(vals)
        valores_globales = list(set(valores_globales))

        # Controles en document.xml
        root = ET.fromstring(xml_files["document"])
        controles = []

        for sdt in root.findall(".//w:sdt", NS_ALL):

            props = sdt.find("w:sdtPr", NS_ALL)
            if props is None:
                continue

            alias_node = props.find("w:alias", NS_ALL)
            tag_node = props.find("w:tag", NS_ALL)
            id_node = props.find("w:id", NS_ALL)

            alias = alias_node.get(f"{{{NS_ALL['w']}}}val") if alias_node is not None else None
            tag = tag_node.get(f"{{{NS_ALL['w']}}}val") if tag_node is not None else None
            cid = id_node.get(f"{{{NS_ALL['w']}}}val") if id_node is not None else None

            tipo = None
            valores_cc = []

            drop = props.find("w:dropDownList", NS_ALL)
            combo = props.find("w:comboBox", NS_ALL)

            if drop is not None or combo is not None:
                tipo = "dropDownList" if drop is not None else "comboBox"

                # valores directos dentro del control
                for entry in props.findall(".//w:listEntry", NS_ALL):
                    val = entry.get(f"{{{NS_ALL['w']}}}value")
                    if val:
                        valores_cc.append(val)

                for entry in props.findall(".//w:listItem", NS_ALL):
                    val = entry.get(f"{{{NS_ALL['w']}}}value")
                    if val:
                        valores_cc.append(val)

                if not valores_cc:
                    valores_cc = valores_globales

                valores_cc = list(set(valores_cc))

            elif props.find("w:checkbox", NS_ALL) is not None:
                tipo = "checkbox"
            elif props.find("w:date", NS_ALL) is not None:
                tipo = "date"
            else:
                tipo = "text"

            # Inferencia lógica
            if alias and isinstance(alias, str) and "fecha" in alias.lower():
                inferred_type = "date"
            elif tipo == "dropDownList":
                inferred_type = "dropdown"
            elif tipo == "checkbox":
                inferred_type = "bool"
            else:
                inferred_type = "text"

            controles.append({
                "alias": alias,
                "tag": tag,
                "id": cid,
                "tipo": tipo,
                "valores": valores_cc or None,
                "inferred_type": inferred_type
            })

        return controles


# ==========================================================
# LECTURA VALOR/FÓRMULA DE CELDA
# ==========================================================
def get_cell_formula_or_value(cell, sheet_xml=None):
    # ArrayFormula
    if isinstance(cell.value, ArrayFormula):
        if hasattr(cell.value, "text"):
            f = cell.value.text.strip()
            return "=" + f if not f.startswith("=") else f

    # Fórmula simple
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return cell.value

    # Búsqueda directa en XML
    if sheet_xml:
        patt = rf'<c[^>]*r="{cell.coordinate}"[^>]*>.*?<f[^>]*>(.*?)</f>'
        m = re.search(patt, sheet_xml, flags=re.DOTALL)
        if m:
            f = m.group(1).strip()
            return "=" + f if not f.startswith("=") else f

    return cell.value


# ==========================================================
# EXTRACCIÓN DE EXCEL EMBEBIDOS
# ==========================================================
def extract_embedded_excel(docx_path):
    with zipfile.ZipFile(docx_path) as z:

        excel_files = [
            f for f in z.namelist()
            if f.startswith("word/embeddings/") and f.endswith(".xlsx")
        ]

        results = []

        for ef in excel_files:

            content = z.read(ef)

            # Cargamos workbook
            wb = load_workbook(BytesIO(content), data_only=False)

            # Leemos XMLs internos (hojas + styles.xml)
            sheet_xmls = {}
            styles_xml_text = None
            with zipfile.ZipFile(BytesIO(content)) as xz:
                for item in xz.namelist():
                    if item.startswith("xl/worksheets/sheet") and item.endswith(".xml"):
                        key = item.split("/")[-1].replace(".xml", "")  # sheet1, sheet2...
                        sheet_xmls[key] = xz.read(item).decode("utf-8", "ignore")
                    if item == "xl/styles.xml":
                        styles_xml_text = xz.read(item).decode("utf-8", "ignore")

            styles_resolver = StylesResolver(styles_xml_text)

            excel_data = {
                "excel": ef,
                "tablas": [],
                "nombres_definidos": []
            }

            # Nombres definidos
            for name, defn in wb.defined_names.items():
                for ws_name, cell_ref in list(defn.destinations):
                    ws = wb[ws_name.replace("'", "")]
                    xml = sheet_xmls.get(f"sheet{ws._id}", "")

                    try:
                        celdas = ws[cell_ref]
                        if hasattr(celdas, "value"):
                            val = get_cell_formula_or_value(celdas, xml)
                        else:
                            val = [
                                [get_cell_formula_or_value(c, xml) for c in fila]
                                for fila in celdas
                            ]
                    except Exception:
                        val = None

                    excel_data["nombres_definidos"].append({
                        "nombre": name,
                        "ref": f"{ws_name}!{cell_ref}",
                        "valor": val
                    })

            # Tablas (ListObjects)
            for ws in wb.worksheets:
                xml = sheet_xmls.get(f"sheet{ws._id}", "")
                for tbl in ws._tables.values():

                    ref = tbl.ref
                    rows = []
                    cell_styles = {}

                    excel_range = ws[ref]

                    # Primera fila (encabezados) → estilos completos celda por celda
                    if excel_range:
                        first_row = excel_range[0]
                        for c in first_row:
                            cell_styles[c.coordinate] = get_complete_cell_format(
                                c, sheet_xml=xml, styles_resolver=styles_resolver
                            )

                    # Valores de toda la tabla
                    for row in excel_range:
                        rows.append([get_cell_formula_or_value(c, xml) for c in row])

                    excel_data["tablas"].append({
                        "worksheet": ws.title,
                        "tabla": tbl.name,
                        "rango": ref,
                        "columnas": rows[0] if rows else [],
                        "registros": rows[1:] if len(rows) > 1 else [],
                        "cell_styles": cell_styles,  # estilos 100% de la primera fila
                    })

            results.append(excel_data)

        return results


# ==========================================================
# MAIN
# ==========================================================
if __name__ == "__main__":

    docx_path = "REPORTE.docx"

    print("\n=== CONTROLES DE CONTENIDO DETECTADOS ===")
    for cc in extract_content_controls(docx_path):
        print(cc)

    print("\n=== EXCEL EMBEBIDOS DETECTADOS ===")
    for ex in extract_embedded_excel(docx_path):
        print("\nArchivo:", ex["excel"])

        print("\n--- Nombres Definidos ---")
        for nd in ex["nombres_definidos"]:
            print(nd)

        print("\n--- Tablas ---")
        for t in ex["tablas"]:
            print(t)
